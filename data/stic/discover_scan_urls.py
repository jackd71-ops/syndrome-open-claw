#!/usr/bin/env python3
"""
Scan URL auto-discovery via Google UK search.

Reads Scan LN codes from Retailer_IDs sheet, searches Google UK for each
LN code (site:scan.co.uk), extracts the first matching product URL, and
writes it back to the "Scan URL" column. Skips rows that already have
a Scan URL filled in.

Run once after Scan LN codes have been manually populated in the sheet.
Usage: python3 discover_scan_urls.py [--dry-run] [--limit N]
"""

import re
import sys
import time
import random
import argparse
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright

TEMPLATE_PATH = "/opt/openclaw/data/general/Retailer_Template.xlsx"
LOG_PATH      = "/opt/openclaw/data/stic/discover_scan.log"
ONEDRIVE_DEST = "onedrive:Documents/Retail Review/Retailer_Template.xlsx"

# Delay between Google searches (seconds) — keep polite to avoid rate limiting
MIN_DELAY = 6
MAX_DELAY = 12


def log(msg):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")


def load_ln_codes():
    """Read all rows from Retailer_IDs sheet that have a Scan LN but no Scan URL."""
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb["Retailer_IDs"]

    # Discover column positions from header row
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    scan_ln_col  = headers.get("Scan LN")
    scan_url_col = headers.get("Scan URL")
    prod_id_col  = headers.get("Product ID") or 1
    model_col    = headers.get("Model No") or 3

    if not scan_ln_col:
        log("ERROR: 'Scan LN' column not found in Retailer_IDs sheet")
        wb.close()
        return [], None, None

    if not scan_url_col:
        log("ERROR: 'Scan URL' column not found in Retailer_IDs sheet")
        wb.close()
        return [], None, None

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod_id  = str(row[prod_id_col - 1]).strip()  if row[prod_id_col - 1]  else None
        ln_code  = str(row[scan_ln_col - 1]).strip()  if row[scan_ln_col - 1]  else None
        scan_url = str(row[scan_url_col - 1]).strip() if row[scan_url_col - 1] else None
        model    = str(row[model_col - 1]).strip()    if row[model_col - 1]    else ""

        if not prod_id or not ln_code:
            continue
        if ln_code.lower() in ("none", ""):
            continue
        if scan_url and scan_url.lower() not in ("none", ""):
            continue  # already populated — skip

        rows.append({
            "product_id": prod_id,
            "ln_code":    ln_code,
            "model":      model,
        })

    wb.close()
    log(f"Found {len(rows)} products with Scan LN but no Scan URL")
    return rows, scan_ln_col, scan_url_col


def google_scan_url(page, ln_code):
    """
    Search Google UK for <LN_CODE> site:scan.co.uk and return the first
    scan.co.uk/products/... URL found, or None.
    """
    query = f"{ln_code} site:scan.co.uk"
    search_url = f"https://www.google.co.uk/search?q={query}&hl=en-GB&gl=GB&num=5"

    try:
        page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
        time.sleep(random.uniform(3, 5))
    except Exception as e:
        log(f"  Navigation error for {ln_code}: {e}")
        return None

    # Check for CAPTCHA / consent wall
    page_text = page.inner_text("body")[:400].lower()
    if "before you continue" in page_text or "i'm not a robot" in page_text:
        log(f"  ⚠️  Google consent/CAPTCHA encountered for {ln_code}")
        return None

    # Extract scan.co.uk product URLs from search result links
    urls = page.evaluate(r"""() => {
        const found = [];
        for (const a of document.querySelectorAll('a[href]')) {
            const h = a.href || '';
            // Google redirects: /url?q=https://www.scan.co.uk/products/...
            const m = h.match(/[?&]q=(https?:\/\/(?:www\.)?scan\.co\.uk\/products\/[^&]+)/);
            if (m) {
                found.push(decodeURIComponent(m[1]));
                continue;
            }
            // Direct links (sometimes appear in AMP or featured snippets)
            if (/scan\.co\.uk\/products\//.test(h) && !h.includes('google.')) {
                found.push(h);
            }
        }
        // Deduplicate preserving order
        return [...new Set(found)];
    }""")

    if urls:
        # Prefer URLs that contain the LN code itself
        ln_lower = ln_code.lower()
        for u in urls:
            if ln_lower in u.lower():
                return u
        return urls[0]

    # Fallback: parse raw text for scan.co.uk/products/ URLs
    raw = page.content()
    matches = re.findall(r'https?://(?:www\.)?scan\.co\.uk/products/[^\s"\'<>]+', raw)
    if matches:
        # Clean up any trailing junk
        cleaned = [re.sub(r'["\'>]+$', '', m) for m in matches]
        cleaned = [u for u in cleaned if u.startswith("http")]
        if cleaned:
            return cleaned[0]

    return None


def write_urls(matches, scan_url_col, dry_run=False):
    """Write discovered Scan URLs into the Retailer_IDs sheet."""
    if not matches:
        log("No URLs to write.")
        return 0

    if dry_run:
        log(f"[DRY RUN] Would write {len(matches)} URLs")
        for pid, url in matches.items():
            log(f"  {pid} → {url}")
        return len(matches)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Retailer_IDs"]

    # Build product_id → row number mapping
    id_to_row = {}
    for row in ws.iter_rows(min_row=2):
        pid = str(row[0].value).strip() if row[0].value else None
        if pid:
            id_to_row[pid] = row[0].row

    written = 0
    for product_id, url in matches.items():
        row_num = id_to_row.get(product_id)
        if row_num:
            existing = ws.cell(row=row_num, column=scan_url_col).value
            if not existing:
                ws.cell(row=row_num, column=scan_url_col).value = url
                written += 1

    wb.save(TEMPLATE_PATH)
    log(f"Wrote {written} Scan URLs to Retailer_IDs sheet.")
    return written


def sync_onedrive():
    r = subprocess.run(
        ["rclone", "copyto", TEMPLATE_PATH, ONEDRIVE_DEST],
        capture_output=True, text=True
    )
    if r.returncode == 0:
        log("OneDrive sync: OK")
    else:
        log(f"OneDrive sync ERROR: {r.stderr.strip()}")


def main():
    parser = argparse.ArgumentParser(description="Discover Scan product URLs via Google UK")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print URLs without writing to sheet")
    parser.add_argument("--limit", type=int, default=0,
                        help="Process at most N products (0 = all)")
    args = parser.parse_args()

    log("=" * 65)
    log("Starting Scan URL discovery via Google UK")
    if args.dry_run:
        log("DRY RUN MODE — no writes")

    rows, scan_ln_col, scan_url_col = load_ln_codes()
    if not rows:
        log("Nothing to process. Exiting.")
        return

    if args.limit and args.limit > 0:
        rows = rows[:args.limit]
        log(f"Limiting to {args.limit} products")

    matches = {}   # product_id → scan URL
    failed  = []   # LN codes that returned nothing

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                  "--disable-dev-shm-usage"]
        )
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-GB",
            timezone_id="Europe/London",
            extra_http_headers={
                "Accept-Language": "en-GB,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "DNT": "1",
            }
        )
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-GB', 'en']});
            Object.defineProperty(navigator, 'plugins',   {get: () => [1, 2, 3]});
            window.chrome = {runtime: {}};
        """)
        page = context.new_page()

        for i, row in enumerate(rows, 1):
            ln   = row["ln_code"]
            pid  = row["product_id"]
            model = row["model"]
            log(f"[{i}/{len(rows)}] {ln}  ({model[:40]})")

            url = google_scan_url(page, ln)

            if url:
                log(f"  ✅ {url}")
                matches[pid] = url
            else:
                log(f"  ❌ No URL found")
                failed.append(ln)

            # Polite delay between searches
            if i < len(rows):
                delay = random.uniform(MIN_DELAY, MAX_DELAY)
                time.sleep(delay)

        browser.close()

    log(f"\n--- Results ---")
    log(f"Found:   {len(matches)}/{len(rows)}")
    log(f"Missing: {len(failed)}")
    if failed:
        log("Unresolved LN codes:")
        for ln in failed:
            log(f"  - {ln}")

    written = write_urls(matches, scan_url_col, dry_run=args.dry_run)

    if written and not args.dry_run:
        sync_onedrive()

    log("Done.")


if __name__ == "__main__":
    main()
