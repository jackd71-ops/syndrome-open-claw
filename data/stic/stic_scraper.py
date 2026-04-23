#!/usr/bin/env python3
"""
STIC (Stock In The Channel) price scraper for OpenClaw.
Searches each product by model number, scrapes distributor prices/stock,
writes results to a dated sheet in the XLSX template.

Usage:
  python3 stic_scraper.py --batch 1   # products 1-260 (9:30am run)
  python3 stic_scraper.py --batch 2   # products 261-520 (1:00pm run)
  python3 stic_scraper.py --test      # first 20 products only
"""

import argparse
import json
import os
import random
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, numbers
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

GBP_FORMAT  = '£#,##0.00'
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# ── Paths ────────────────────────────────────────────────────────────────────
MASTER_PATH   = "/opt/openclaw/data/general/STIC Template.xlsx"   # never written to
CACHE_PATH    = "/opt/openclaw/data/stic/url_cache.json"
PROGRESS_PATH = "/opt/openclaw/data/stic/progress_{date}.json"
SESSION_PATH  = "/opt/openclaw/data/stic/session.json"
LOG_PATH      = "/opt/openclaw/logs/stic.log"

def get_monthly_path() -> str:
    """Returns e.g. /opt/openclaw/data/general/STIC_2026-04.xlsx"""
    month = datetime.now().strftime("%Y-%m")
    return f"/opt/openclaw/data/general/STIC_{month}.xlsx"

# ── STIC config ───────────────────────────────────────────────────────────────
STIC_BASE     = "https://www.stockinthechannel.co.uk"
STIC_LOGIN    = "https://www.stockinthechannel.co.uk/Account/Login"
STIC_SEARCH   = "https://www.stockinthechannel.co.uk/Search?q={query}"

# Credentials — stored in secrets.json
SECRETS_PATH  = "/opt/openclaw/secrets.json"

# OneDrive destination (rclone remote path)
ONEDRIVE_DEST = "onedrive:Documents/STIC"

# Telegram
TELEGRAM_CHAT_ID = "1163684840"

# ── Target distributors (must match column headers in template exactly) ───────
DISTRIBUTORS = [
    "TD Synnex UK",
    "VIP",
    "Westcoast",
    "Target",
    "M2M Direct",
]

# Partial match aliases — STIC may show slightly different names
DISTRIBUTOR_ALIASES = {
    "TD Synnex UK":  ["td synnex", "tdsynnex", "synnex"],
    "VIP":           ["vip", "vip computers", "vip distribution"],
    "Westcoast":     ["westcoast", "west coast"],
    "Target":        ["target", "target components"],
    "M2M Direct":    ["m2m", "m2m direct"],
}

# ── Timing ────────────────────────────────────────────────────────────────────
DELAY_MIN         = 4
DELAY_MAX         = 12
LONG_PAUSE_EVERY  = random.randint(30, 50)   # products between long pauses
LONG_PAUSE_MIN    = 20
LONG_PAUSE_MAX    = 60

# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")

# ── Secrets ───────────────────────────────────────────────────────────────────
def get_credentials():
    with open(SECRETS_PATH) as f:
        secrets = json.load(f)
    return secrets.get("STIC_USERNAME"), secrets.get("STIC_PASSWORD")

# ── URL cache ─────────────────────────────────────────────────────────────────
def load_cache() -> dict:
    if Path(CACHE_PATH).exists():
        with open(CACHE_PATH) as f:
            return json.load(f)
    return {}

def save_cache(cache: dict):
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)

# ── Progress tracking ─────────────────────────────────────────────────────────
def load_progress(date_str: str) -> set:
    path = PROGRESS_PATH.format(date=date_str)
    if Path(path).exists():
        with open(path) as f:
            return set(json.load(f))
    return set()

def save_progress(date_str: str, completed: set):
    path = PROGRESS_PATH.format(date=date_str)
    with open(path, "w") as f:
        json.dump(list(completed), f)

# ── Read products from template ───────────────────────────────────────────────
def ensure_monthly_file(monthly_path: str):
    """Create this month's working file from the master template if it doesn't exist."""
    import shutil
    if Path(monthly_path).exists():
        return
    shutil.copy2(MASTER_PATH, monthly_path)
    log(f"Created new monthly file: {Path(monthly_path).name}")

def count_products() -> int:
    """Count total products in master template."""
    wb = load_workbook(MASTER_PATH, read_only=True)
    ws = wb.worksheets[0]
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            count += 1
    wb.close()
    return count

def batch_ranges(total: int) -> list[tuple[int, int]]:
    """Split total products into 3 roughly equal ranges."""
    size = total // 3
    remainder = total % 3
    ranges = []
    start = 1
    for i in range(3):
        extra = 1 if i < remainder else 0
        end = start + size + extra - 1
        ranges.append((start, end))
        start = end + 1
    return ranges

def read_products(start: int, end: int) -> list:
    """Read products from master template. Returns list of dicts."""
    wb = load_workbook(MASTER_PATH, read_only=True)
    # Master sheet is first sheet ("Templae")
    ws = wb.worksheets[0]
    products = []
    row_num = 0
    for row in ws.iter_rows(min_row=3, values_only=True):  # skip 2 header rows
        if row[0] is None:
            continue
        row_num += 1
        if row_num < start:
            continue
        if row_num > end:
            break
        eol = row[7] if len(row) > 7 else None
        if eol:
            continue  # skip end-of-life products
        products.append({
            "row_num":       row_num,
            "product_id":    row[0],
            "description":   str(row[1]).strip() if row[1] else "",
            "model_no":      str(row[2]).strip() if row[2] else "",
            "manufacturer":  str(row[3]).strip() if row[3] else "",
            "product_group": str(row[4]).strip() if row[4] else None,
            "chipset":       str(row[5]).strip() if len(row) > 5 and row[5] else None,
            "ean":           str(row[6]).strip() if len(row) > 6 and row[6] else None,
        })
    wb.close()
    return products

# ── Ensure dated sheet exists ─────────────────────────────────────────────────
def ensure_dated_sheet(monthly_path: str, date_str: str):
    """Add a dated sheet to this month's file if it doesn't exist yet."""
    from copy import copy

    wb = load_workbook(monthly_path)
    master = wb.worksheets[0]

    if date_str in wb.sheetnames:
        log(f"Sheet '{date_str}' already exists.")
        wb.close()
        return

    new_ws = wb.copy_worksheet(master)
    new_ws.title = date_str

    # Clear price/qty columns (J onwards), keep product info (A-H) and spacer (I)
    for row in new_ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column >= 10:
                cell.value = None

    # Ensure column T header = "Total Stock"
    for r in (1, 2):
        cell = new_ws.cell(row=r, column=20)
        if not cell.value:
            cell.value = "Total Stock"

    wb.save(monthly_path)
    wb.close()
    log(f"Created sheet '{date_str}' in {Path(monthly_path).name}.")

# ── Write result to sheet ─────────────────────────────────────────────────────
def write_result(monthly_path: str, date_str: str, row_num: int, distributor_data: dict, status: str = None):
    """Write scraped data to the dated sheet. Row num is 1-indexed product row."""
    wb = load_workbook(monthly_path)
    ws = wb[date_str]

    # Excel row = row_num + 2 (2 header rows)
    excel_row = row_num + 2

    # Column mapping: G=TD Synnex Price, H=TD Synnex Qty, I=VIP Price, J=VIP Qty,
    # K=Westcoast Price, L=Westcoast Qty, M=Target Price, N=Target Qty,
    # O=M2M Price, P=M2M Qty
    col_map = {
        "TD Synnex UK": (10, 11),  # J, K
        "VIP":          (12, 13),  # L, M
        "Westcoast":    (14, 15),  # N, O
        "Target":       (16, 17),  # P, Q
        "M2M Direct":   (18, 19),  # R, S
    }

    if status == "FAILED_MATCH":
        # Mark column A with failed match indicator
        ws.cell(row=excel_row, column=1).value = f"FAILED_MATCH: {ws.cell(row=excel_row, column=1).value}"
    elif status == "NOT_LISTED":
        pass  # Leave blank — distributor doesn't list this product
    else:
        price_cells = {}  # dist_name -> cell column, for green highlight
        for dist_name, (price_col, qty_col) in col_map.items():
            if dist_name in distributor_data:
                price, qty = distributor_data[dist_name]
                if price is not None:
                    c = ws.cell(row=excel_row, column=price_col)
                    c.value = price
                    c.number_format = GBP_FORMAT
                    price_cells[dist_name] = price_col
                if qty is not None:
                    ws.cell(row=excel_row, column=qty_col).value = qty

        # Green highlight on cheapest price cell
        if price_cells:
            prices_with_stock = {
                d: distributor_data[d][0]
                for d in price_cells
                if distributor_data[d][0] is not None and (distributor_data[d][1] or 0) > 0
            }
            # Fall back to all prices if none have stock
            compare = prices_with_stock if prices_with_stock else {
                d: distributor_data[d][0] for d in price_cells if distributor_data[d][0] is not None
            }
            if compare:
                cheapest_dist = min(compare, key=compare.get)
                ws.cell(row=excel_row, column=price_cells[cheapest_dist]).fill = GREEN_FILL

        # Column T (20): total stock across all monitored distributors
        qty_cols = [qty_col for _, (_, qty_col) in col_map.items()]
        qty_letters = [get_column_letter(c) for c in qty_cols]
        sum_formula = "=" + "+".join(f"IFERROR({l}{excel_row},0)" for l in qty_letters)
        ws.cell(row=excel_row, column=20).value = sum_formula

    wb.save(monthly_path)
    wb.close()

# ── Data bleed detection ─────────────────────────────────────────────────────
def check_data_bleed(iso_date: str) -> list[dict]:
    """
    Find SKU pairs where 3+ distributors share identical non-null price+qty on
    the same date — strong signal that one SKU picked up another's scraped data.
    Returns list of dicts: {product_id, model_no, matched_to, matched_model, matching_rows}
    """
    import sqlite3
    try:
        db = sqlite3.connect(_DB_PATH)
        db.row_factory = sqlite3.Row
        rows = db.execute("""
            SELECT a.product_id, a.model_no, b.product_id AS matched_to,
                   b.model_no AS matched_model, COUNT(*) AS matching_rows
            FROM stic_prices a
            JOIN stic_prices b
              ON b.date=a.date AND b.distributor=a.distributor
             AND b.product_id != a.product_id
             AND a.price IS NOT NULL AND b.price IS NOT NULL
             AND a.qty   IS NOT NULL AND b.qty   IS NOT NULL
             AND a.price = b.price   AND a.qty   = b.qty
            WHERE a.date=? AND a.product_id < b.product_id
            GROUP BY a.product_id, b.product_id
            HAVING matching_rows >= 3
            ORDER BY matching_rows DESC, a.product_id
        """, (iso_date,)).fetchall()
        db.close()
        return [dict(r) for r in rows]
    except Exception as e:
        log(f"  Data bleed check error: {e}")
        return []


# ── Write result to SQLite ────────────────────────────────────────────────────
_DB_PATH = "/opt/openclaw/data/analytics/prices.db"

_DIST_DB_NAME = {
    "TD Synnex UK": "TD Synnex",
    "VIP":          "VIP",
    "Westcoast":    "Westcoast",
    "Target":       "Target",
    "M2M Direct":   "M2M Direct",
}

def write_to_db(date_str: str, product: dict, distributor_data: dict):
    """Write scraped distributor prices to SQLite. Never raises — logs on failure."""
    import sqlite3
    try:
        # Convert DD-MM-YYYY to YYYY-MM-DD for DB consistency
        d, m, y = date_str.split("-")
        iso_date = f"{y}-{m}-{d}"

        product_id    = product["product_id"]
        model_no      = product["model_no"]
        manufacturer  = product["manufacturer"]
        product_group = product.get("product_group")
        chipset       = product.get("chipset")

        db = sqlite3.connect(_DB_PATH)
        db.execute("PRAGMA journal_mode=WAL")
        for dist_name in _DIST_DB_NAME:
            db_dist = _DIST_DB_NAME[dist_name]
            price, qty = distributor_data.get(dist_name, (None, None))
            db.execute(
                """INSERT OR IGNORE INTO stic_prices
                   (date, product_id, model_no, manufacturer, product_group,
                    chipset, distributor, price, qty)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (iso_date, product_id, model_no, manufacturer, product_group,
                 chipset, db_dist, price, qty),
            )
        db.commit()
        db.close()
    except Exception as e:
        log(f"  DB write error (non-fatal): {e}")

# ── Browser: login ────────────────────────────────────────────────────────────
def login(page, username: str, password: str) -> bool:
    log("Logging into STIC...")
    page.goto(STIC_LOGIN, wait_until="domcontentloaded")
    time.sleep(random.uniform(2, 4))

    try:
        # Fill username
        page.wait_for_selector('input[type="email"]', timeout=10000)
        page.fill('input[type="email"]', username)
        time.sleep(random.uniform(0.8, 1.5))

        # Fill password
        page.fill('input[type="password"]', password)
        time.sleep(random.uniform(0.8, 1.5))

        # Screenshot pre-submit for debug
        page.screenshot(path="/opt/openclaw/data/stic/pre_submit.png")

        # Submit by pressing Enter on password field — more reliable than button click
        page.press('input[type="password"]', "Enter")
        time.sleep(8)  # wait for redirect
        page.wait_for_load_state("domcontentloaded")
        time.sleep(random.uniform(2, 3))

        # Dump URL and key page indicators for diagnosis
        current_url = page.url
        content = page.content().lower()
        log(f"Post-login URL: {current_url}")
        log(f"Page has 'logout': {'logout' in content}")
        log(f"Page has 'sign out': {'sign out' in content}")
        log(f"Page has 'my account': {'my account' in content}")
        log(f"Page has 'register': {'register' in content}")

        # Save screenshot for inspection
        page.screenshot(path="/opt/openclaw/data/stic/login_debug.png")
        log("Screenshot saved to /opt/openclaw/data/stic/login_debug.png")

        if "logout" in content or "sign out" in content or "my account" in content:
            log("Login successful.")
            return True
        else:
            log(f"Login FAILED — still seeing guest content.")
            return False
    except Exception as e:
        log(f"Login error: {e}")
        return False

# ── Browser: search and scrape in one step ───────────────────────────────────
def search_and_scrape(page, model_no: str, cache: dict, product_id: str = None, manufacturer: str = "") -> dict | None:
    """
    Search for model_no on STIC, scrape the distributor table from the search
    results page (prices are shown inline — no need to visit product page).
    If multiple result cards are found, fall back to searching by product_id.

    Returns dict: { "Distributor Name": (price_float_or_none, qty_int_or_none) }
    Returns None on page load failure, {} if no matching distributors found.
    """
    search_url = STIC_SEARCH.format(query=model_no.replace(" ", "+"))
    log(f"  Searching: {model_no}")

    try:
        page.goto(search_url, wait_until="domcontentloaded")
        time.sleep(random.uniform(2, 4))

        # Human-like scroll
        page.mouse.wheel(0, random.randint(100, 400))
        time.sleep(random.uniform(0.5, 1.5))

        # Check we got results
        content = page.content()
        if "0 Results Found" in content or "No results" in content.lower():
            log(f"  No results for: {model_no}")
            return {}

        # Find the product card whose SKU matches our model_no, then scrape only that table.
        # Falls back to brand match. Never uses first table blindly — returns match_type='none'
        # if neither check passes, so we write nulls rather than bleed another SKU's data.
        result = page.evaluate("""
            ([modelNo, brand]) => {
                const tables = document.querySelectorAll('table');
                if (!tables.length) return { match_type: 'none', rows: [] };

                const brandLower = (brand || '').toLowerCase();

                function cardText(table) {
                    let el = table.parentElement;
                    for (let i = 0; i < 12; i++) {
                        if (!el) break;
                        el = el.parentElement;
                    }
                    return '';
                }

                function getCardRoot(table) {
                    let el = table.parentElement;
                    for (let i = 0; i < 12; i++) {
                        if (!el) break;
                        const text = el.innerText || '';
                        if (text.includes('SKU:') || text.includes('Brand:') || text.includes('Manufacturer:')) return el;
                        el = el.parentElement;
                    }
                    return table.parentElement;
                }

                let targetTable = null;
                let matchType = 'none';

                // Pass 1: exact SKU match
                for (const table of tables) {
                    let el = table.parentElement;
                    for (let i = 0; i < 12; i++) {
                        if (!el) break;
                        const text = el.innerText || '';
                        if (text.includes('SKU: ' + modelNo)) {
                            targetTable = table;
                            matchType = 'sku';
                            break;
                        }
                        el = el.parentElement;
                    }
                    if (targetTable) break;
                }

                // Pass 2: brand match (only if brand provided and no SKU match)
                if (!targetTable && brandLower) {
                    for (const table of tables) {
                        let el = table.parentElement;
                        for (let i = 0; i < 12; i++) {
                            if (!el) break;
                            const text = (el.innerText || '').toLowerCase();
                            if (text.includes(brandLower)) {
                                targetTable = table;
                                matchType = 'brand';
                                break;
                            }
                            el = el.parentElement;
                        }
                        if (targetTable) break;
                    }
                }

                if (!targetTable) return { match_type: 'none', rows: [] };

                const rows = [];
                targetTable.querySelectorAll('tr').forEach((row, idx) => {
                    if (idx === 0) return;
                    const cells = row.querySelectorAll('td');
                    if (cells.length < 2) return;
                    const distName = cells[0].innerText.trim();
                    const stockText = cells.length >= 3 ? cells[cells.length - 2].innerText.trim() : '';
                    const priceText = cells[cells.length - 1].innerText.trim();
                    if (distName) rows.push({
                        distributor: distName,
                        stock: stockText,
                        price: priceText,
                        allCells: Array.from(cells).map(c => c.innerText.trim())
                    });
                });

                return { match_type: matchType, rows };
            }
        """, [model_no, manufacturer])

        match_type = result.get("match_type", "none") if isinstance(result, dict) else "none"
        raw_rows   = result.get("rows", [])         if isinstance(result, dict) else (result or [])

        if match_type == "none":
            log(f"  No SKU or brand match on results page — will try product detail fallback.")
            raw_rows = []
        elif match_type == "brand":
            log(f"  WARNING: no SKU match — accepted on brand match only ({manufacturer}). Verify manually.")

        result = raw_rows

        # If no table found, retry via product ID:
        # 1. Search by ID to get results page
        # 2. Click through to the first product detail page
        # 3. Scrape the 6-column Trade Prices table there
        if not result and product_id:
            # Build fallback query: "asus 121284" if manufacturer known, else just "121284"
            fallback_query = f"{manufacturer} {product_id}".strip() if manufacturer else str(product_id)
            log(f"  No table found — retrying via product detail page (query: {fallback_query})")
            id_url = STIC_SEARCH.format(query=fallback_query.replace(" ", "+"))
            page.goto(id_url, wait_until="domcontentloaded")
            time.sleep(random.uniform(2, 3))

            # Find and click the first product link
            product_links = page.query_selector_all('a[href*="/Product/"]')
            if product_links:
                product_href = product_links[0].get_attribute("href")
                if product_href and not product_href.startswith("http"):
                    product_href = STIC_BASE + product_href
                log(f"  Following product link: {product_href}")
                page.goto(product_href, wait_until="domcontentloaded")
                time.sleep(random.uniform(2, 4))
                page.mouse.wheel(0, random.randint(200, 400))
                time.sleep(random.uniform(1, 2))

                # Scrape 6-col Trade Prices table: Distributor|Product|SKU|Stock|Updated|Price
                result = page.evaluate("""
                    () => {
                        const data = [];
                        const tables = document.querySelectorAll('table');
                        // Find the trade prices table — has 6 columns
                        let targetTable = null;
                        for (const t of tables) {
                            const headerRow = t.querySelector('tr');
                            if (headerRow) {
                                const text = headerRow.innerText.toLowerCase();
                                if (text.includes('distributor') || text.includes('stock') || text.includes('price')) {
                                    targetTable = t;
                                    break;
                                }
                            }
                        }
                        if (!targetTable) return data;

                        const rows = targetTable.querySelectorAll('tr');
                        rows.forEach((row, idx) => {
                            if (idx === 0) return; // skip header
                            const cells = row.querySelectorAll('td');
                            if (cells.length < 4) return;

                            // Distributor: innerText or img alt
                            let distName = cells[0].innerText.trim();
                            if (!distName) {
                                const img = cells[0].querySelector('img');
                                if (img) distName = img.alt || img.title || '';
                            }

                            // 6-col layout: Distributor|Product|SKU|Stock|Updated|Price
                            // 4-col layout: Distributor|Product|Stock|Price
                            let stockText, priceText;
                            if (cells.length >= 6) {
                                stockText = cells[3].innerText.trim();
                                priceText = cells[5].innerText.trim();
                            } else {
                                stockText = cells[cells.length - 2].innerText.trim();
                                priceText = cells[cells.length - 1].innerText.trim();
                            }

                            if (distName) data.push({
                                distributor: distName,
                                stock: stockText,
                                price: priceText,
                                allCells: Array.from(cells).map(c => c.innerText.trim())
                            });
                        });
                        return data;
                    }
                """)
                if result:
                    # Validate brand on the detail page before accepting
                    page_text = page.inner_text("body").lower()
                    if manufacturer and manufacturer.lower() not in page_text:
                        log(f"  VALIDATION FAILED: brand '{manufacturer}' not found on detail page — discarding data.")
                        result = []
                    else:
                        log(f"  Product detail page fallback succeeded (brand validated).")
                else:
                    log(f"  Product detail page also found no table.")
            else:
                log(f"  No product links found for ID {product_id}.")

        if not result:
            return {}

        cache[model_no] = search_url

        # Parse into { dist_name: (price, qty) }
        parsed = {}
        for row in result:
            dist_raw = row.get("distributor", "").lower().strip()
            if not dist_raw:
                continue

            # Match to our target distributors
            matched_dist = None
            for dist_name, aliases in DISTRIBUTOR_ALIASES.items():
                if any(alias in dist_raw for alias in aliases):
                    matched_dist = dist_name
                    break

            if not matched_dist:
                continue

            all_cells = row.get("allCells", [])
            log(f"  Matched {matched_dist}: cells={all_cells}")

            # Stock — second to last cell
            stock_text = row.get("stock", "").replace(",", "").strip()
            qty = None
            try:
                qty = int(stock_text)
            except (ValueError, TypeError):
                qty = 0 if stock_text in ("0", "", "-", "n/a") else None

            # Price — last cell, expect "£12.34"
            price_text = row.get("price", "").replace("£", "").replace(",", "").strip()
            price = None
            try:
                parsed_price = float(price_text)
                price = parsed_price if parsed_price > 0 else None
            except (ValueError, TypeError):
                price = None

            # If distributor already seen, keep the entry with highest qty (and its price)
            if matched_dist in parsed:
                existing_price, existing_qty = parsed[matched_dist]
                if (qty or 0) > (existing_qty or 0):
                    parsed[matched_dist] = (price, qty)
                # else keep existing — it has higher stock
            else:
                parsed[matched_dist] = (price, qty)
            log(f"  {matched_dist}: price={parsed[matched_dist][0]}, qty={parsed[matched_dist][1]}")

        return parsed

    except PlaywrightTimeout:
        log(f"  Timeout searching for: {model_no}")
        return None
    except Exception as e:
        log(f"  Search error for {model_no}: {e}")
        return None

# ── OneDrive sync ────────────────────────────────────────────────────────────
def sync_template_from_onedrive() -> bool:
    """Pull the latest master template from OneDrive before each run."""
    import subprocess
    log("Syncing master template from OneDrive…")
    result = subprocess.run(
        ["rclone", "copy", f"{ONEDRIVE_DEST}/STIC Template.xlsx", str(Path(MASTER_PATH).parent)],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        log("Master template updated from OneDrive.")
        return True
    else:
        log(f"Template sync WARNING (using local copy): {result.stderr.strip()}")
        return False


def sync_to_onedrive(monthly_path: str) -> bool:
    """Copy this month's STIC Excel file to OneDrive via rclone."""
    import subprocess
    log(f"Syncing {Path(monthly_path).name} → {ONEDRIVE_DEST}/")
    result = subprocess.run(
        ["rclone", "copy", monthly_path, ONEDRIVE_DEST, "--progress"],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        log("OneDrive sync complete.")
        return True
    else:
        log(f"OneDrive sync FAILED: {result.stderr.strip()}")
        return False

# ── Telegram notification ─────────────────────────────────────────────────────
def send_telegram(message: str):
    import requests as req
    with open(SECRETS_PATH) as f:
        token = json.load(f).get("TELEGRAM_TOKEN")
    if not token:
        log("No TELEGRAM_TOKEN in secrets — skipping notification.")
        return
    try:
        resp = req.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"},
            timeout=10,
        )
        if resp.status_code == 200:
            log("Telegram notification sent.")
        else:
            log(f"Telegram error: {resp.status_code} {resp.text}")
    except Exception as e:
        log(f"Telegram exception: {e}")

# ── Human-like mouse movement ─────────────────────────────────────────────────
def random_mouse_move(page):
    """Move mouse to a random position to appear human."""
    x = random.randint(200, 1200)
    y = random.randint(200, 700)
    page.mouse.move(x, y)

# ── Main run ──────────────────────────────────────────────────────────────────
def run(start: int, end: int, date_str: str, is_final: bool = False):
    username, password = get_credentials()
    if not username or not password:
        log("ERROR: STIC_USERNAME or STIC_PASSWORD not found in secrets.json")
        sys.exit(1)

    monthly_path = get_monthly_path()
    cache = load_cache()
    completed = load_progress(date_str)

    sync_template_from_onedrive()

    log(f"Starting STIC scrape: products {start}–{end}, date={date_str}")
    log(f"Monthly file: {Path(monthly_path).name}")
    log(f"URL cache: {len(cache)} entries. Already completed today: {len(completed)}")

    # Ensure this month's file and today's sheet exist
    ensure_monthly_file(monthly_path)
    ensure_dated_sheet(monthly_path, date_str)

    products = read_products(start, end)
    log(f"Loaded {len(products)} products for this batch.")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
            ]
        )

        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="en-GB",
            timezone_id="Europe/London",
        )

        # Add stealth script to mask automation fingerprints
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-GB', 'en'] });
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3] });
            window.chrome = { runtime: {} };
        """)

        page = context.new_page()

        # Login
        if not login(page, username, password):
            log("Login failed — aborting.")
            browser.close()
            sys.exit(1)

        products_since_pause = 0
        long_pause_every = random.randint(30, 50)

        for product in products:
            row_num      = product["row_num"]
            model_no     = product["model_no"]
            prod_id      = product["product_id"]
            manufacturer = product["manufacturer"]

            # Skip if already done today
            if str(row_num) in completed:
                log(f"Skipping {model_no} (row {row_num}) — already done today.")
                continue

            log(f"\n[{row_num}/{end}] {model_no} (ID: {prod_id})")

            # Random mouse movement
            random_mouse_move(page)

            # Search and scrape distributor prices from results page
            dist_data = search_and_scrape(page, model_no, cache, product_id=str(prod_id), manufacturer=manufacturer)

            if dist_data is None:
                log(f"  Page load failed — skipping (will retry next run).")
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
                continue

            if not dist_data and model_no not in cache:
                log(f"  FAILED MATCH — no result found for: {model_no}")
                write_result(monthly_path, date_str, row_num, {}, status="FAILED_MATCH")
                completed.add(str(row_num))
                save_progress(date_str, completed)
                save_cache(cache)
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
                continue

            # Write to Excel and SQLite
            write_result(monthly_path, date_str, row_num, dist_data)
            write_to_db(date_str, product, dist_data)

            completed.add(str(row_num))
            save_progress(date_str, completed)
            save_cache(cache)  # Save cache incrementally

            products_since_pause += 1

            # Long pause every 30-50 products
            if products_since_pause >= long_pause_every:
                pause = random.uniform(LONG_PAUSE_MIN, LONG_PAUSE_MAX)
                log(f"  Long pause: {pause:.0f}s")
                time.sleep(pause)
                products_since_pause = 0
                long_pause_every = random.randint(30, 50)  # reset
            else:
                delay = random.uniform(DELAY_MIN, DELAY_MAX)
                log(f"  Delay: {delay:.1f}s")
                time.sleep(delay)

        browser.close()

    failed = [p["model_no"] for p in products if str(p["row_num"]) not in completed]
    log(f"\nBatch complete. {len(completed)} products done today. {len(failed)} skipped/failed.")

    # Sync to OneDrive after every batch
    sync_ok = sync_to_onedrive(monthly_path)

    # Telegram notification — only on final batch
    if is_final:
        total_today = len(load_progress(date_str))
        status_icon = "✅" if sync_ok else "⚠️"

        # Convert DD-MM-YYYY to YYYY-MM-DD for DB query
        d, m, y = date_str.split("-")
        iso_date = f"{y}-{m}-{d}"
        bleed_suspects = check_data_bleed(iso_date)
        log(f"Data bleed check: {len(bleed_suspects)} suspect pairs found.")

        bleed_section = ""
        if bleed_suspects:
            lines = [f"⚠️ <b>Possible data bleed ({len(bleed_suspects)} pairs) — please verify:</b>"]
            for s in bleed_suspects:
                lines.append(
                    f"  • {s['product_id']} {s['model_no']} ↔ {s['matched_to']} {s['matched_model']}"
                    f" ({s['matching_rows']} distis match)"
                )
            bleed_section = "\n" + "\n".join(lines)
        else:
            bleed_section = "\n✅ No data bleed suspects found."

        msg = (
            f"{status_icon} <b>STIC scrape complete</b> — {date_str}\n\n"
            f"✔️ Total scraped today: {total_today} products\n"
            f"❌ This batch failed/skipped: {len(failed)}\n"
            f"📁 {Path(monthly_path).name} → OneDrive {'synced' if sync_ok else 'SYNC FAILED'}\n"
            f"🕐 Finished: {datetime.now().strftime('%H:%M')}"
            f"{bleed_section}"
        )
        send_telegram(msg)

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=int, choices=[1, 2, 3], help="1/2/3 — split dynamically from template count")
    parser.add_argument("--test",  action="store_true",          help="Test run: first 20 products only")
    parser.add_argument("--start", type=int,                     help="Custom start row")
    parser.add_argument("--end",   type=int,                     help="Custom end row")
    args = parser.parse_args()

    date_str = datetime.now().strftime("%d-%m-%Y")

    if args.test:
        run(1, 20, date_str, is_final=False)
    elif args.batch in (1, 2, 3):
        total = count_products()
        ranges = batch_ranges(total)
        log(f"Template has {total} products — batch ranges: {ranges}")
        start, end = ranges[args.batch - 1]
        is_final = (args.batch == 3)
        run(start, end, date_str, is_final=is_final)
    elif args.start and args.end:
        run(args.start, args.end, date_str, is_final=True)
    else:
        parser.print_help()
