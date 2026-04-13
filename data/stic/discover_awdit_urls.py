#!/usr/bin/env python3
"""
AWD-IT URL auto-discovery.
Browses AWD-IT category pages to build a product name → URL lookup,
then matches against our 492 products and writes URLs into Retailer_IDs sheet.

AWD-IT is a Magento store — category pages list products with names and URLs.
"""

import json
import re
import time
import random
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

TEMPLATE_PATH = "/opt/openclaw/data/general/Retailer_Template.xlsx"
LOG_PATH      = "/opt/openclaw/data/stic/discover_awdit.log"

# AWD-IT category pages most likely to contain our products
# (motherboards, GPUs, cases, etc. — adjust if needed)
CATEGORIES = [
    "https://www.awd-it.co.uk/components/motherboards.html",
    "https://www.awd-it.co.uk/components/motherboards/intel-motherboards.html",
    "https://www.awd-it.co.uk/components/motherboards/amd-motherboards.html",
    "https://www.awd-it.co.uk/components/graphics-cards.html",
    "https://www.awd-it.co.uk/components/graphics-cards/nvidia-graphics-cards.html",
    "https://www.awd-it.co.uk/components/graphics-cards/amd-graphics-cards.html",
    "https://www.awd-it.co.uk/components.html",
    "https://www.awd-it.co.uk/networking.html",
    "https://www.awd-it.co.uk/components/storage.html",
    "https://www.awd-it.co.uk/components/memory.html",
]

def log(msg):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")

def slugify(text):
    """Convert product name to URL slug for matching."""
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def load_products():
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb.worksheets[0]
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        products.append({
            "product_id":   str(row[0]).strip(),
            "model_no":     str(row[2]).strip() if row[2] else "",
            "manufacturer": str(row[3]).strip() if row[3] else "",
        })
    wb.close()
    return products

def is_match(product_name, model_no, manufacturer):
    """Check if an AWD-IT product name matches our model."""
    name_l  = product_name.lower()
    model_l = model_no.lower()
    mfr_l   = manufacturer.lower().split()[0] if manufacturer else ""

    # All significant tokens from model number must appear in the product name
    tokens = [t for t in re.split(r'[\s\-/]+', model_l) if len(t) >= 2]
    if not tokens:
        return False
    match_count = sum(1 for t in tokens if t in name_l)
    # Require manufacturer match + majority of model tokens
    mfr_match = mfr_l in name_l if mfr_l else True
    return mfr_match and match_count >= max(2, len(tokens) * 0.7)

def get_all_pages(page, category_url):
    """Get all product items from a category, handling pagination."""
    all_items = []
    url = category_url
    page_num = 1

    while url:
        log(f"  Fetching page {page_num}: {url}")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            time.sleep(random.uniform(4, 7))
        except Exception as e:
            log(f"  Error loading page: {e}")
            break

        # Check if blocked
        body = page.inner_text("body")[:200].lower()
        if "cloudflare" in body or "just a moment" in body or "blocked" in body:
            log(f"  Blocked on: {url}")
            break

        # Extract product items
        items = page.evaluate("""() => {
            const results = [];
            const cards = document.querySelectorAll('.product-item-info, .product-item, li.product-item');
            for (const card of cards) {
                const link = card.querySelector('a.product-item-link, .product-item-link, a[title]');
                if (link) {
                    results.push({
                        name: (link.getAttribute('title') || link.innerText || '').trim(),
                        url:  link.href || ''
                    });
                }
            }
            return results;
        }""")

        new_items = [i for i in items if i['url'] and i['name'] and 'awd-it.co.uk' in i['url']]
        all_items.extend(new_items)
        log(f"  Found {len(new_items)} products on this page (total: {len(all_items)})")

        if not new_items:
            break

        # Check for next page
        next_url = page.evaluate("""() => {
            const next = document.querySelector('a.action.next, [aria-label="Next"], .pages-item-next a');
            return next ? next.href : null;
        }""")

        if next_url and next_url != url:
            url = next_url
            page_num += 1
        else:
            break

    return all_items

def write_urls_to_sheet(matches):
    """Write discovered URLs into Retailer_IDs sheet."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Retailer_IDs"]

    # Find AWD-IT URL column
    headers = {ws.cell(row=1, column=i).value: i for i in range(1, 15)}
    awdit_col = headers.get("AWD-IT URL")
    if not awdit_col:
        log("ERROR: AWD-IT URL column not found in Retailer_IDs sheet")
        return 0

    # Build product_id → row mapping
    id_to_row = {}
    for row in ws.iter_rows(min_row=2):
        pid = str(row[0].value).strip() if row[0].value else None
        if pid:
            id_to_row[pid] = row[0].row

    written = 0
    for product_id, url in matches.items():
        row_num = id_to_row.get(product_id)
        if row_num:
            existing = ws.cell(row=row_num, column=awdit_col).value
            if not existing:  # don't overwrite manual entries
                ws.cell(row=row_num, column=awdit_col).value = url
                written += 1

    wb.save(TEMPLATE_PATH)
    log(f"Wrote {written} URLs to Retailer_IDs sheet.")
    return written

def main():
    log("Starting AWD-IT URL discovery")
    products = load_products()
    log(f"Loaded {len(products)} products to match")

    # Build catalog from AWD-IT category pages
    catalog = {}  # name → url

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled","--disable-dev-shm-usage"]
        )
        context = browser.new_context(
            viewport={"width":1366,"height":768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="en-GB", timezone_id="Europe/London",
            extra_http_headers={"Accept-Language":"en-GB,en;q=0.9","DNT":"1"}
        )
        context.add_init_script("""
            Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
            Object.defineProperty(navigator,'languages',{get:()=>['en-GB','en']});
            Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});
            window.chrome={runtime:{}};
        """)
        page = context.new_page()

        # Phase 1: navigate AWD-IT to find real category URLs
        log("\n--- Phase 1: Discovering category URLs from navigation ---")
        page.goto("https://www.awd-it.co.uk", wait_until="domcontentloaded", timeout=25000)
        time.sleep(4)

        nav_links = page.evaluate("""() => {
            return [...document.querySelectorAll('nav a, .navigation a, #store\\.menu a')]
                .map(a => ({text: a.innerText.trim(), href: a.href}))
                .filter(a => a.href.includes('awd-it.co.uk') && a.text.length > 2 && a.text.length < 50)
                .slice(0, 40);
        }""")
        log(f"Nav links found: {len(nav_links)}")
        for link in nav_links:
            log(f"  {link['text'][:40]:40s} → {link['href'][:60]}")

        # Combine discovered nav links with our predefined categories
        component_urls = list(set(
            CATEGORIES +
            [l['href'] for l in nav_links if any(
                kw in l['text'].lower() for kw in
                ['motherboard', 'graphics', 'gpu', 'component', 'storage', 'memory', 'networking', 'monitor']
            )]
        ))
        log(f"\n--- Phase 2: Scraping {len(component_urls)} category pages ---")

        for cat_url in component_urls:
            log(f"\nCategory: {cat_url}")
            items = get_all_pages(page, cat_url)
            for item in items:
                catalog[item['name']] = item['url']
            time.sleep(random.uniform(3, 6))

        browser.close()

    log(f"\n--- Phase 3: Matching {len(catalog)} catalog items against {len(products)} products ---")

    # Save catalog for inspection
    catalog_path = "/opt/openclaw/data/stic/awdit_catalog.json"
    with open(catalog_path, "w") as f:
        json.dump(catalog, f, indent=2)
    log(f"Catalog saved to {catalog_path}")

    matches = {}
    no_match = []

    for product in products:
        model_no     = product["model_no"]
        manufacturer = product["manufacturer"]
        product_id   = product["product_id"]

        best_match = None
        for cat_name, cat_url in catalog.items():
            if is_match(cat_name, model_no, manufacturer):
                best_match = (cat_name, cat_url)
                break

        if best_match:
            log(f"  ✅ {model_no:35s} → {best_match[0][:50]}")
            matches[product_id] = best_match[1]
        else:
            no_match.append(model_no)

    log(f"\n--- Results ---")
    log(f"Matched:     {len(matches)}/{len(products)}")
    log(f"No match:    {len(no_match)}")

    if no_match:
        log(f"\nUnmatched products:")
        for m in no_match[:20]:
            log(f"  - {m}")

    if matches:
        written = write_urls_to_sheet(matches)
        log(f"Written to sheet: {written}")

        # Sync template to OneDrive
        import subprocess
        r = subprocess.run(
            ["rclone", "copyto", TEMPLATE_PATH,
             "onedrive:Documents/Retail Review/Retailer_Template.xlsx"],
            capture_output=True, text=True
        )
        log(f"OneDrive sync: {'OK' if r.returncode == 0 else r.stderr}")

if __name__ == "__main__":
    main()
