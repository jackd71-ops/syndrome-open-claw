#!/usr/bin/env python3
"""
Backfill stic_prices table from STIC_2026-04.xlsx.
Reads every dated sheet, inserts one row per product×distributor.
Safe to re-run — uses INSERT OR IGNORE.
Usage: python3 backfill_stic.py
"""
import re
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = "/opt/openclaw/data/general/STIC_2026-04.xlsx"
DB_PATH   = "/opt/openclaw/data/analytics/prices.db"

# Col indices (0-based): A=0 product_id, B=1 desc, C=2 model_no,
# D=3 manufacturer, E=4 product_group, F=5 spacer,
# G=6 TD price, H=7 TD qty, I=8 VIP price, J=9 VIP qty,
# K=10 Westcoast price, L=11 Westcoast qty,
# M=12 Target price, N=13 Target qty,
# O=14 M2M price, P=15 M2M qty, Q=16 Total Stock (formula — skip)

DIST_COLS = [
    ("TD Synnex",  6,  7),
    ("VIP",        8,  9),
    ("Westcoast",  10, 11),
    ("Target",     12, 13),
    ("M2M Direct", 14, 15),
]

DATE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")  # DD-MM-YYYY


def sheet_to_iso(name: str) -> str | None:
    m = DATE_RE.match(name)
    if not m:
        return None
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def to_int(v) -> int | None:
    if v is None:
        return None
    try:
        return int(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def backfill():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")

    total_inserted = 0
    total_skipped  = 0

    for sheet_name in wb.sheetnames:
        iso_date = sheet_to_iso(sheet_name)
        if not iso_date:
            print(f"  Skipping sheet '{sheet_name}' (not a date)")
            continue

        ws = wb[sheet_name]
        rows_inserted = 0
        rows_skipped  = 0

        for row in ws.iter_rows(min_row=3, values_only=True):
            product_id = row[0]
            if product_id is None:
                continue
            try:
                product_id = int(product_id)
            except (ValueError, TypeError):
                continue

            model_no      = str(row[2]).strip() if row[2] else ""
            manufacturer  = str(row[3]).strip() if row[3] else ""
            product_group = str(row[4]).strip() if row[4] else None

            for dist_name, price_col, qty_col in DIST_COLS:
                price = to_float(row[price_col])
                qty   = to_int(row[qty_col])

                cur = db.execute(
                    """INSERT OR IGNORE INTO stic_prices
                       (date, product_id, model_no, manufacturer, product_group,
                        distributor, price, qty)
                       VALUES (?,?,?,?,?,?,?,?)""",
                    (iso_date, product_id, model_no, manufacturer, product_group,
                     dist_name, price, qty),
                )
                if cur.rowcount:
                    rows_inserted += 1
                else:
                    rows_skipped += 1

        db.commit()
        print(f"  {sheet_name} ({iso_date}): +{rows_inserted} inserted, {rows_skipped} skipped")
        total_inserted += rows_inserted
        total_skipped  += rows_skipped

    wb.close()
    db.close()
    print(f"\nDone. Total inserted: {total_inserted}, skipped: {total_skipped}")


if __name__ == "__main__":
    backfill()
