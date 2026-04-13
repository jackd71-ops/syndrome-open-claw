#!/usr/bin/env python3
"""Probe direct product page URLs and price selectors for each ID-based retailer."""
import time, re
from playwright.sync_api import sync_playwright

# Test with real codes from the sheet
TESTS = [
    ("Amazon",  "https://www.amazon.co.uk/dp/B0BXFBN121"),
    ("Currys",  "https://www.currys.co.uk/products/x/690644.html"),
    ("Very",    "https://www.very.co.uk/x/690644.prd"),          # placeholder — need real format
    ("Argos",   "https://www.argos.co.uk/product/WL5645006/"),   # test "WL " prefix handling
    ("Argos2",  "https://www.argos.co.uk/product/5645006/"),     # test without WL
]

# Separate Very test — need to figure out URL from a real SKU
# Very SKU from sheet appears blank for row 1-2, let's find a populated one

def get_very_sample():
    from openpyxl import load_workbook
    wb = load_workbook("/opt/openclaw/data/general/Retailer_Template.xlsx", read_only=True)
    ws = wb["Retailer_IDs"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6]:  # Very SKU is col G (index 6)
            print(f"Very SKU sample: model={row[2]!r} sku={row[6]!r}")
            return str(row[6]).strip()
    return None

def get_argos_sample():
    from openpyxl import load_workbook
    wb = load_workbook("/opt/openclaw/data/general/Retailer_Template.xlsx", read_only=True)
    ws = wb["Retailer_IDs"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[7]:  # Argos SKU is col H (index 7)
            print(f"Argos SKU sample: model={row[2]!r} sku={row[7]!r}")
            return str(row[7]).strip()
    return None

very_sku  = get_very_sample()
argos_sku = get_argos_sample()
print(f"\nVery SKU: {very_sku!r}")
print(f"Argos SKU: {argos_sku!r}\n")

def probe(page, name, url):
    print(f"\n{'='*65}")
    print(f"  {name}")
    print(f"  {url}")
    print('='*65)
    try:
        resp = page.goto(url, wait_until="domcontentloaded", timeout=25000)
        time.sleep(5)
        final_url = page.url
        print(f"  Final URL: {final_url}")
        print(f"  Status: {resp.status if resp else 'N/A'}")

        page.screenshot(path=f"/opt/openclaw/data/stic/probe_direct_{name}.png")

        body = page.inner_text("body")[:200].lower().replace('\n', ' ')
        print(f"  Body: {body[:150]}")

        # Find £ price elements
        pound_els = page.evaluate("""() => {
            const results = [];
            for (const el of document.querySelectorAll('*')) {
                const direct = [...el.childNodes]
                    .filter(n => n.nodeType === 3)
                    .map(n => n.textContent).join('').trim();
                if (direct.match(/£\\s*\\d+\\.\\d{2}/) && direct.length < 60) {
                    results.push({
                        tag: el.tagName,
                        cls: (el.className||'').toString().substring(0,80),
                        id:  (el.id||'').substring(0,30),
                        text: direct.substring(0,50)
                    });
                    if (results.length >= 8) break;
                }
            }
            return results;
        }""")

        if pound_els:
            print(f"\n  £ price elements:")
            for e in pound_els:
                ident = f"#{e['id']}" if e['id'] else f".{e['cls'][:60]}"
                print(f"    <{e['tag']}> {ident}")
                print(f"         {e['text']!r}")
        else:
            print("  ❌ No £ price elements found")

        # JSON-LD
        json_ld = page.evaluate("""() => {
            return [...document.querySelectorAll('script[type="application/ld+json"]')]
                   .map(s => s.textContent)
                   .filter(t => t.includes('price') || t.includes('Price'))
                   .map(t => t.substring(0, 400));
        }""")
        if json_ld:
            print(f"\n  JSON-LD:")
            for j in json_ld[:1]:
                print(f"    {j[:300]}")

    except Exception as e:
        print(f"  ERROR: {e}")

# Build test list with real SKUs
tests = [
    ("Amazon",  "https://www.amazon.co.uk/dp/B0BXFBN121"),
    ("Currys",  "https://www.currys.co.uk/products/x/690644.html"),
]

if very_sku:
    # Very SKUs are typically numeric — URL format: very.co.uk/x/{sku}.prd
    tests.append(("Very", f"https://www.very.co.uk/x/{very_sku}.prd"))

if argos_sku:
    clean_sku = argos_sku.replace("WL ", "").strip()
    tests.append(("Argos_clean",  f"https://www.argos.co.uk/product/{clean_sku}/"))
    tests.append(("Argos_raw",    f"https://www.argos.co.uk/product/{argos_sku.replace(' ','')}/"))

with sync_playwright() as p:
    browser = p.chromium.launch(
        headless=True,
        args=["--no-sandbox","--disable-blink-features=AutomationControlled","--disable-dev-shm-usage"]
    )
    ctx = browser.new_context(
        viewport={"width":1366,"height":768},
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        locale="en-GB", timezone_id="Europe/London",
    )
    ctx.add_init_script("""
        Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
        Object.defineProperty(navigator,'languages',{get:()=>['en-GB','en']});
        Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});
        window.chrome={runtime:{}};
    """)
    page = ctx.new_page()
    for name, url in tests:
        probe(page, name, url)
        time.sleep(4)
    browser.close()
