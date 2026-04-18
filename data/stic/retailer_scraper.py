#!/usr/bin/env python3
"""
Retailer price scraper for OpenClaw.
Uses direct product IDs (ASINs / Currys SKUs) from Retailer_IDs sheet
for guaranteed accuracy; falls back to model-number search where no ID exists.

Retailer status (2026-04):
  ✅ Amazon UK    — dp/{ASIN} direct page (homepage warm-up for cookies)
  ✅ Currys       — search by SKU → direct product redirect
  ✅ Overclockers — OCUK code search via camoufox (code in col M of template)
  ❌ Argos        — 403 on all URLs (Akamai)
  ⚠️  Scan         — Cloudflare blocks direct; URLs discovered via Google UK search
  ❌ Box          — full JS render
  ✅ CCL Online   — direct product URL → JSON-LD price
  ✅ AWD-IT       — direct product URL → JSON-LD price (URLs auto-discovered)
  ❌ Very         — 403 on all URLs

Pre-flight discovery (runs automatically before batch 1):
  • AWD-IT: matches new products against cached category catalog; re-scrapes
    AWD-IT category pages only when unmatched products remain.
  • Scan: for any product with a Scan LN code but no Scan URL, searches
    Google UK (site:scan.co.uk) and saves the discovered URL.

Usage:
  python3 retailer_scraper.py --batch 1|2|3
  python3 retailer_scraper.py --test
  python3 retailer_scraper.py --start 1 --end 25
  python3 retailer_scraper.py --discover        # run only the pre-flight phase
"""

import argparse
import json
import re
import subprocess
import random
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import sys
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# ── Paths ─────────────────────────────────────────────────────────────────────
TEMPLATE_PATH      = "/opt/openclaw/data/general/Retailer_Template.xlsx"
OUTPUT_DIR         = "/opt/openclaw/data/general"
PROGRESS_PATH      = "/opt/openclaw/data/stic/retailer_progress_{date}.json"
LOG_PATH           = "/opt/openclaw/logs/retailer.log"
SECRETS_PATH       = "/opt/openclaw/secrets.json"
ONEDRIVE_DEST      = "onedrive:Documents/Retail Review/"
AWDIT_CATALOG_PATH = "/opt/openclaw/data/stic/awdit_catalog.json"
SCAN_SCRAPE_PATH   = "/opt/openclaw/data/stic/scan_scrape.py"
VERY_SCRAPE_PATH   = "/opt/openclaw/data/stic/very_scrape.py"

# ── AWD-IT category pages for URL discovery ───────────────────────────────────
AWDIT_CATEGORIES = [
    "https://www.awd-it.co.uk/components/motherboards.html",
    "https://www.awd-it.co.uk/components/motherboards/intel-motherboards.html",
    "https://www.awd-it.co.uk/components/motherboards/amd-motherboards.html",
    "https://www.awd-it.co.uk/components/graphics-cards.html",
    "https://www.awd-it.co.uk/components/graphics-cards/nvidia-graphics-cards.html",
    "https://www.awd-it.co.uk/components/graphics-cards/amd-graphics-cards.html",
    "https://www.awd-it.co.uk/components.html",
    "https://www.awd-it.co.uk/networking.html",
    "https://www.awd-it.co.uk/components/storage.html",
    "https://www.awd-it.co.uk/components/memory.html",
]

# Polite delay between Google searches (seconds)
GOOGLE_DELAY_MIN = 6
GOOGLE_DELAY_MAX = 12

# ── Retailer definitions ───────────────────────────────────────────────────────
# id_col : column name in Retailer_IDs sheet (None = search-only, no IDs)
# blocked: skip entirely, write "—"
RETAILERS = [
    {"name": "Amazon UK",    "col": 10, "id_col": "Amazon ASIN", "blocked": False},
    {"name": "Currys",       "col": 11, "id_col": "Currys SKU",  "blocked": False},
    {"name": "Argos",        "col": 12, "id_col": "ARGOS  SKU",  "blocked": True, "reason": "403 Akamai"},
    {"name": "Scan",         "col": 13, "id_col": "Scan URL",    "blocked": False},
    {"name": "Overclockers", "col": 14, "id_col": None,          "blocked": False},
    {"name": "Box",          "col": 15, "id_col": None,          "blocked": True, "reason": "JS render"},
    {"name": "CCL Online",   "col": 16, "id_col": "CCL URL",     "blocked": False},
    {"name": "AWD-IT",       "col": 17, "id_col": "AWD-IT URL",  "blocked": False},
    {"name": "Very",         "col": 18, "id_col": "Very URL",    "blocked": False},
]

# ── Timing ────────────────────────────────────────────────────────────────────
DELAY_MIN      = 6
DELAY_MAX      = 14
LONG_PAUSE_MIN = 30
LONG_PAUSE_MAX = 90

# ── Styles ────────────────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", fgColor="FF0000")
RED_FONT    = Font(color="FFFFFF", bold=True)
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
GREY_FONT   = Font(color="808080", italic=True)
ORANGE_FILL = PatternFill("solid", fgColor="FFA500")
ORANGE_FONT = Font(color="FFFFFF", bold=True)

# ── Logging ───────────────────────────────────────────────────────────────────
def log(msg):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")

def get_secrets():
    with open(SECRETS_PATH) as f:
        return json.load(f)

def get_output_path():
    return f"{OUTPUT_DIR}/Retailer_Results_{datetime.now().strftime('%Y-%m')}.xlsx"

# ── Progress ──────────────────────────────────────────────────────────────────
def load_progress(date_str):
    path = PROGRESS_PATH.format(date=date_str)
    if Path(path).exists():
        with open(path) as f:
            return set(json.load(f))
    return set()

def save_progress(date_str, completed):
    with open(PROGRESS_PATH.format(date=date_str), "w") as f:
        json.dump(list(completed), f)

# ── Dynamic batch split ───────────────────────────────────────────────────────
def count_products():
    wb    = load_workbook(TEMPLATE_PATH, read_only=True)
    ws    = wb.worksheets[0]
    count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if row[0] is not None)
    wb.close()
    return count

def batch_ranges(total):
    size, rem = total // 3, total % 3
    ranges, start = [], 1
    for i in range(3):
        end = start + size + (1 if i < rem else 0) - 1
        ranges.append((start, end))
        start = end + 1
    return ranges

# ── Load Retailer_IDs sheet ───────────────────────────────────────────────────
def load_retailer_ids():
    """Returns dict: {product_id: {retailer_id_col_name: code}}"""
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    if "Retailer_IDs" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Retailer_IDs"]
    # Row 1 = headers; col A = Product ID
    headers = [ws.cell(row=1, column=i).value for i in range(1, 20)]
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        prod_id = row[0]
        if prod_id is None:
            continue
        codes = {}
        for i, hdr in enumerate(headers):
            if hdr and i >= 4:  # cols E onwards are retailer IDs
                val = row[i]
                if val:
                    codes[hdr] = str(val).strip()
        mapping[str(prod_id).strip()] = codes
    wb.close()
    return mapping

# ── Read products ─────────────────────────────────────────────────────────────
def read_products(start, end):
    wb      = load_workbook(TEMPLATE_PATH, read_only=True)
    ws      = wb.worksheets[0]
    products, row_num = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        row_num += 1
        if row_num < start:
            continue
        if row_num > end:
            break
        msrp = None
        try:
            msrp = float(row[7]) if row[7] is not None else None
        except (ValueError, TypeError):
            pass
        products.append({
            "row_num":      row_num,
            "product_id":   str(row[0]).strip(),
            "model_no":     str(row[2]).strip() if row[2] else "",
            "manufacturer": str(row[3]).strip() if row[3] else "",
            "msrp":         msrp,
        })
    wb.close()
    return products

# ── Ensure monthly output file ────────────────────────────────────────────────
def ensure_output_file(date_str):
    import shutil
    out = get_output_path()
    if not Path(out).exists():
        shutil.copy(TEMPLATE_PATH, out)
        log(f"Created monthly file: {Path(out).name}")
    wb = load_workbook(out)
    if date_str not in wb.sheetnames:
        master = wb.worksheets[0]
        ws     = wb.copy_worksheet(master)
        ws.title = date_str
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column >= 10:
                    cell.value = None
                    cell.fill  = PatternFill(fill_type=None)
        wb.save(out)
        log(f"Created sheet '{date_str}'")
    wb.close()

# ── Write one cell ────────────────────────────────────────────────────────────
def write_cell(date_str, excel_row, col, price, msrp, status=None):
    out  = get_output_path()
    wb   = load_workbook(out)
    ws   = wb[date_str]
    cell = ws.cell(row=excel_row, column=col)
    if status == "BLOCKED":
        cell.value = "—"
        cell.fill  = GREY_FILL
        cell.font  = GREY_FONT
    elif status == "OUT_OF_STOCK":
        cell.value = "OOS"
        cell.fill  = ORANGE_FILL
        cell.font  = ORANGE_FONT
    elif price is not None:
        cell.value         = price
        cell.number_format = '£#,##0.00'
        if msrp and price < msrp:
            cell.fill = RED_FILL
            cell.font = RED_FONT
    # status == "NOT_FOUND" → leave blank
    wb.save(out)
    wb.close()

# ── Price parser ──────────────────────────────────────────────────────────────
def parse_price(text):
    if not text:
        return None
    text = text.replace(",", "").replace("\xa0", "").strip()
    m = re.search(r'£\s*(\d+\.\d{2})', text)
    if m:
        try:
            v = float(m.group(1))
            return v if v > 0.5 else None
        except ValueError:
            pass
    return None

# ── Amazon scraper (direct dp/ page) ─────────────────────────────────────────
amazon_warmed_up = False

def warm_up_amazon(page):
    global amazon_warmed_up
    if amazon_warmed_up:
        return
    log("  [Amazon] Warming up cookies via homepage...")
    page.goto("https://www.amazon.co.uk", wait_until="domcontentloaded", timeout=20000)
    time.sleep(4)
    amazon_warmed_up = True

def scrape_amazon(page, asin):
    warm_up_amazon(page)
    url = f"https://www.amazon.co.uk/dp/{asin}"
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    time.sleep(random.uniform(4, 7))
    for sel in [".a-price .a-offscreen", "#apex_offerDisplay_desktop .a-price .a-offscreen"]:
        try:
            el = page.query_selector(sel)
            if el:
                p = parse_price(el.inner_text())
                if p:
                    return p
        except Exception:
            pass
    return None

# ── Currys scraper (search by SKU → product page redirect) ───────────────────
def scrape_currys(page, sku):
    url = f"https://www.currys.co.uk/search?q={sku}"
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    time.sleep(random.uniform(5, 9))
    for sel in [".product-price", "[class*='ProductPrice']", "[class*='price']"]:
        try:
            el = page.query_selector(sel)
            if el:
                p = parse_price(el.inner_text())
                if p:
                    return p
        except Exception:
            pass
    return None

# ── CCL scraper (direct product URL → JSON-LD price) ─────────────────────────
def scrape_ccl(page, url):
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    time.sleep(random.uniform(4, 7))
    price = page.evaluate("""() => {
        for (const s of document.querySelectorAll('script[type="application/ld+json"]')) {
            try {
                const d = JSON.parse(s.textContent);
                if (d.offers && d.offers.price) return parseFloat(d.offers.price);
                if (d['@graph']) {
                    for (const item of d['@graph']) {
                        if (item.offers && item.offers.price) return parseFloat(item.offers.price);
                    }
                }
            } catch(e) {}
        }
        return null;
    }""")
    return float(price) if price and price > 0 else None

# ── AWD-IT scraper (direct product URL → JSON-LD price) ──────────────────────
def scrape_awdit(page, url):
    page.goto(url, wait_until="domcontentloaded", timeout=25000)
    time.sleep(random.uniform(4, 7))
    price = page.evaluate("""() => {
        for (const s of document.querySelectorAll('script[type="application/ld+json"]')) {
            try {
                const d = JSON.parse(s.textContent);
                if (d['@type'] === 'Product' && d.offers && d.offers.price) {
                    return parseFloat(d.offers.price);
                }
            } catch(e) {}
        }
        return null;
    }""")
    return float(price) if price and price > 0 else None

# ── Scan scraper (patchright + Xvfb subprocess) ──────────────────────────────
def scrape_scan(url):
    try:
        result = subprocess.run(
            ["xvfb-run", "--auto-servernum", "/usr/bin/python3", SCAN_SCRAPE_PATH, url],
            capture_output=True, text=True, timeout=60
        )
        out = result.stdout.strip()
        if out and out != "NOT_FOUND":
            return float(out)
    except (subprocess.TimeoutExpired, ValueError, Exception):
        pass
    return None


# ── Very scraper (patchright + Xvfb subprocess) ───────────────────────────────
def scrape_very(url):
    try:
        result = subprocess.run(
            ["xvfb-run", "--auto-servernum", "/usr/bin/python3", VERY_SCRAPE_PATH, url],
            capture_output=True, text=True, timeout=90
        )
        out = result.stdout.strip()
        if out and out != "NOT_FOUND":
            return float(out)
    except (subprocess.TimeoutExpired, ValueError, Exception):
        pass
    return None


# ── Overclockers scraper (camoufox subprocess) ───────────────────────────────
OCUK_SCRAPE_PATH = "/opt/openclaw/data/stic/ocuk_scrape.py"

def scrape_overclockers(code):
    try:
        result = subprocess.run(
            ["/usr/bin/python3", OCUK_SCRAPE_PATH, code],
            capture_output=True, text=True, timeout=60
        )
        out = result.stdout.strip()
        if out and out != "NOT_FOUND":
            return float(out)
    except (subprocess.TimeoutExpired, ValueError, Exception):
        pass
    return None

# ── Main scrape per product ───────────────────────────────────────────────────
def scrape_product(page, product, retailer, id_codes):
    name     = retailer["name"]
    id_col   = retailer.get("id_col")

    # Scan: patchright + Xvfb subprocess, uses stored Scan URL
    if name == "Scan":
        url = id_codes.get(product["product_id"], {}).get("Scan URL")
        if not url:
            return None, "NOT_STOCKED"
        try:
            price = scrape_scan(url)
            return (price, None) if price else (None, "NOT_FOUND")
        except Exception as e:
            log(f"    [{name}] ERROR: {e}")
            return None, "ERROR"

    # Very: patchright + Xvfb subprocess, uses stored Very URL
    if name == "Very":
        url = id_codes.get(product["product_id"], {}).get("Very URL")
        sku = id_codes.get(product["product_id"], {}).get("Very SKU")
        if not url:
            return None, "OUT_OF_STOCK" if sku else "NOT_STOCKED"
        try:
            price = scrape_very(url)
            return (price, None) if price else (None, "NOT_FOUND")
        except Exception as e:
            log(f"    [{name}] ERROR: {e}")
            return None, "ERROR"

    # Overclockers: uses OCUK code from product dict, scraped via camoufox subprocess
    if name == "Overclockers":
        code = id_codes.get(product["product_id"], {}).get("OCUK Code")
        if not code:
            return None, "NOT_STOCKED"
        try:
            price = scrape_overclockers(code)
            return (price, None) if price else (None, "NOT_FOUND")
        except Exception as e:
            log(f"    [{name}] ERROR: {e}")
            return None, "ERROR"

    # All other retailers: use id_codes from Retailer_IDs sheet
    prod_id = product["product_id"]
    code    = id_codes.get(prod_id, {}).get(id_col) if id_col else None

    if id_col and not code:
        return None, "NOT_STOCKED"

    try:
        if name == "Amazon UK":
            price = scrape_amazon(page, code)
        elif name == "Currys":
            price = scrape_currys(page, code)
        elif name == "CCL Online":
            price = scrape_ccl(page, code)
        elif name == "AWD-IT":
            price = scrape_awdit(page, code)
        else:
            return None, "NOT_FOUND"

        return (price, None) if price else (None, "NOT_FOUND")

    except PlaywrightTimeout:
        return None, "TIMEOUT"
    except Exception as e:
        log(f"    [{name}] ERROR: {e}")
        return None, "ERROR"

# ── Telegram ──────────────────────────────────────────────────────────────────
def send_telegram(message):
    import urllib.request
    token = get_secrets().get("TELEGRAM_TOKEN", "")
    if not token:
        return
    url  = f"https://api.telegram.org/bot{token}/sendMessage"
    data = json.dumps({"chat_id": "1163684840", "text": message, "parse_mode": "HTML"}).encode()
    req  = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        urllib.request.urlopen(req, timeout=10)
        log("Telegram sent.")
    except Exception as e:
        log(f"Telegram error: {e}")

# ── OneDrive sync ─────────────────────────────────────────────────────────────
def sync_onedrive():
    """Push current monthly results file to OneDrive."""
    out = get_output_path()
    r   = subprocess.run(["rclone", "copyto", out, f"{ONEDRIVE_DEST}{Path(out).name}"],
                         capture_output=True, text=True)
    ok  = r.returncode == 0
    log("OneDrive sync: OK" if ok else f"OneDrive sync FAILED: {r.stderr}")
    return ok


def sync_template_to_onedrive():
    """Push local Retailer_Template.xlsx to OneDrive (after discovery writes URLs)."""
    r = subprocess.run(
        ["rclone", "copyto", TEMPLATE_PATH, f"{ONEDRIVE_DEST}Retailer_Template.xlsx"],
        capture_output=True, text=True
    )
    ok = r.returncode == 0
    log("Template → OneDrive: OK" if ok else f"Template → OneDrive FAILED: {r.stderr.strip()}")
    return ok


def sync_template_from_onedrive():
    """
    Pull Retailer_Template.xlsx from OneDrive if the remote copy is newer.
    Runs at the start of batch 1 so any codes the user added at work
    (ASINs, LN codes, CCL URLs etc.) are picked up before scraping begins.
    """
    log("Checking OneDrive for updated template...")
    r = subprocess.run(
        ["rclone", "copyto", "--update",
         f"{ONEDRIVE_DEST}Retailer_Template.xlsx", TEMPLATE_PATH],
        capture_output=True, text=True
    )
    if r.returncode == 0:
        log("Template ← OneDrive: OK (pulled if remote was newer)")
    else:
        log(f"Template ← OneDrive FAILED: {r.stderr.strip()} — using local copy")

# ═══════════════════════════════════════════════════════════════════════════════
# PRE-FLIGHT URL DISCOVERY
# ═══════════════════════════════════════════════════════════════════════════════

# ── Helpers: read Retailer_IDs columns ───────────────────────────────────────
def _ids_sheet_headers():
    """Return (ws, header→col_index dict) for Retailer_IDs sheet (1-based)."""
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    ws = wb["Retailer_IDs"]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    wb.close()
    return headers


def load_products_needing_awdit():
    """Return list of {product_id, model_no, manufacturer} rows with no AWD-IT URL."""
    wb      = load_workbook(TEMPLATE_PATH, read_only=True)
    ws      = wb["Retailer_IDs"]
    hdrs    = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    awdit_c = hdrs.get("AWD-IT URL")
    mfr_c   = hdrs.get("Manufacturer", 2)
    model_c = hdrs.get("Model No", 3)
    rows    = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[0]).strip() if row[0] else None
        if not pid:
            continue
        url = str(row[awdit_c - 1]).strip() if awdit_c and row[awdit_c - 1] else ""
        if url and url.lower() not in ("none", ""):
            continue
        rows.append({
            "product_id":   pid,
            "model_no":     str(row[model_c - 1]).strip()  if row[model_c - 1] else "",
            "manufacturer": str(row[mfr_c - 1]).strip()   if row[mfr_c - 1]  else "",
        })
    wb.close()
    return rows


def load_products_needing_scan_url():
    """Return list of {product_id, ln_code, model_no} rows with Scan LN but no Scan URL."""
    wb      = load_workbook(TEMPLATE_PATH, read_only=True)
    ws      = wb["Retailer_IDs"]
    hdrs    = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    ln_c    = hdrs.get("Scan LN")
    url_c   = hdrs.get("Scan URL")
    model_c = hdrs.get("Model No", 3)
    if not ln_c or not url_c:
        wb.close()
        return []
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[0]).strip() if row[0] else None
        if not pid:
            continue
        ln  = str(row[ln_c  - 1]).strip() if row[ln_c  - 1] else ""
        url = str(row[url_c - 1]).strip() if row[url_c - 1] else ""
        if not ln or ln.lower() == "none":
            continue
        if url and url.lower() not in ("none", ""):
            continue
        rows.append({
            "product_id": pid,
            "ln_code":    ln,
            "model_no":   str(row[model_c - 1]).strip() if row[model_c - 1] else "",
        })
    wb.close()
    return rows


def load_products_needing_very_url():
    """Return list of {product_id, sku, model_no} rows with Very SKU but no Very URL."""
    wb      = load_workbook(TEMPLATE_PATH, read_only=True)
    ws      = wb["Retailer_IDs"]
    hdrs    = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    sku_c   = hdrs.get("Very SKU")
    url_c   = hdrs.get("Very URL")
    model_c = hdrs.get("Model No", 3)
    if not sku_c:
        wb.close()
        return []
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[0]).strip() if row[0] else None
        if not pid:
            continue
        sku = str(row[sku_c - 1]).strip() if row[sku_c - 1] else ""
        url = str(row[url_c - 1]).strip() if url_c and row[url_c - 1] else ""
        if not sku or sku.lower() == "none":
            continue
        if url and url.lower() not in ("none", ""):
            continue
        rows.append({
            "product_id": pid,
            "sku":        sku,
            "model_no":   str(row[model_c - 1]).strip() if row[model_c - 1] else "",
        })
    wb.close()
    return rows


def _write_discovered_urls(col_name, id_to_url):
    """Write {product_id: url} into the named column of Retailer_IDs sheet."""
    if not id_to_url:
        return 0
    wb   = load_workbook(TEMPLATE_PATH)
    ws   = wb["Retailer_IDs"]
    hdrs = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    col  = hdrs.get(col_name)
    if not col:
        log(f"[Discovery] ERROR: column '{col_name}' not found in Retailer_IDs")
        wb.close()
        return 0
    id_to_row = {}
    for row in ws.iter_rows(min_row=2):
        pid = str(row[0].value).strip() if row[0].value else None
        if pid:
            id_to_row[pid] = row[0].row
    written = 0
    for pid, url in id_to_url.items():
        row_num = id_to_row.get(pid)
        if row_num:
            existing = ws.cell(row=row_num, column=col).value
            if not existing:
                ws.cell(row=row_num, column=col).value = url
                written += 1
    wb.save(TEMPLATE_PATH)
    log(f"[Discovery] Wrote {written} URLs to '{col_name}' column.")
    return written


# ── AWD-IT: token matching ────────────────────────────────────────────────────
def _is_awdit_match(catalog_name, model_no, manufacturer):
    name_l  = catalog_name.lower()
    model_l = model_no.lower()
    mfr_l   = manufacturer.lower().split()[0] if manufacturer else ""
    tokens  = [t for t in re.split(r'[\s\-/]+', model_l) if len(t) >= 2]
    if not tokens:
        return False
    match_count = sum(1 for t in tokens if t in name_l)
    mfr_match   = (mfr_l in name_l) if mfr_l else True
    return mfr_match and match_count >= max(2, len(tokens) * 0.7)


# ── AWD-IT: scrape one category page (all paginated pages) ───────────────────
def _awdit_scrape_category(page, category_url):
    """Return list of {name, url} across all pagination pages for one category."""
    all_items, url, page_num = [], category_url, 1
    while url:
        log(f"  [AWD-IT] Category page {page_num}: {url}")
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=25000)
            time.sleep(random.uniform(4, 7))
        except Exception as e:
            log(f"  [AWD-IT] Error: {e}")
            break
        body = page.inner_text("body")[:200].lower()
        if any(x in body for x in ("cloudflare", "just a moment", "blocked")):
            log("  [AWD-IT] Blocked — stopping category.")
            break
        items = page.evaluate("""() => {
            const results = [];
            for (const card of document.querySelectorAll(
                    '.product-item-info, .product-item, li.product-item')) {
                const link = card.querySelector(
                    'a.product-item-link, .product-item-link, a[title]');
                if (link) results.push({
                    name: (link.getAttribute('title') || link.innerText || '').trim(),
                    url:  link.href || ''
                });
            }
            return results;
        }""")
        new = [i for i in items if i["url"] and i["name"] and "awd-it.co.uk" in i["url"]]
        all_items.extend(new)
        log(f"  [AWD-IT] {len(new)} products (total: {len(all_items)})")
        if not new:
            break
        next_url = page.evaluate("""() => {
            const n = document.querySelector(
                'a.action.next, [aria-label="Next"], .pages-item-next a');
            return n ? n.href : null;
        }""")
        if next_url and next_url != url:
            url = next_url
            page_num += 1
        else:
            break
    return all_items


# ── AWD-IT discovery: match new products from cache, scrape if needed ─────────
def discover_awdit_urls(page):
    """
    Discover AWD-IT product URLs for products not yet in the Retailer_IDs sheet.
    1. Load cached catalog (JSON) — try to match new products from it first.
    2. Only scrape fresh AWD-IT category pages if unmatched products remain.
    3. Write newly found URLs to 'AWD-IT URL' column and save catalog.
    """
    needs = load_products_needing_awdit()
    if not needs:
        log("[Discovery] AWD-IT: all products already have URLs — skipping.")
        return 0

    log(f"[Discovery] AWD-IT: {len(needs)} products need URLs.")

    # Load existing catalog cache
    catalog = {}
    if Path(AWDIT_CATALOG_PATH).exists():
        try:
            with open(AWDIT_CATALOG_PATH) as f:
                catalog = json.load(f)
            log(f"[Discovery] AWD-IT: loaded {len(catalog)} cached catalog entries.")
        except Exception as e:
            log(f"[Discovery] AWD-IT: cache load error: {e}")

    def try_match(products):
        matched, unmatched = {}, []
        for p in products:
            hit = None
            for cat_name, cat_url in catalog.items():
                if _is_awdit_match(cat_name, p["model_no"], p["manufacturer"]):
                    hit = cat_url
                    break
            if hit:
                matched[p["product_id"]] = hit
                log(f"  [AWD-IT] ✅ {p['model_no'][:40]} (from cache)")
            else:
                unmatched.append(p)
        return matched, unmatched

    matched, still_needs = try_match(needs)

    if still_needs:
        log(f"[Discovery] AWD-IT: {len(still_needs)} still unmatched — scraping category pages...")

        # Discover extra category URLs from AWD-IT navigation
        try:
            page.goto("https://www.awd-it.co.uk", wait_until="domcontentloaded", timeout=25000)
            time.sleep(4)
            nav_links = page.evaluate("""() => {
                return [...document.querySelectorAll('nav a, .navigation a, #store\\.menu a')]
                    .map(a => ({text: a.innerText.trim(), href: a.href}))
                    .filter(a => a.href.includes('awd-it.co.uk') &&
                                 a.text.length > 2 && a.text.length < 50)
                    .slice(0, 50);
            }""")
            extra = [
                l["href"] for l in nav_links
                if any(kw in l["text"].lower() for kw in
                       ["motherboard", "graphics", "gpu", "component",
                        "storage", "memory", "networking", "monitor"])
            ]
            category_urls = list(set(AWDIT_CATEGORIES + extra))
        except Exception as e:
            log(f"[Discovery] AWD-IT: nav error ({e}), using predefined categories.")
            category_urls = AWDIT_CATEGORIES

        for cat_url in category_urls:
            items = _awdit_scrape_category(page, cat_url)
            for item in items:
                catalog[item["name"]] = item["url"]
            time.sleep(random.uniform(3, 6))

        # Save updated catalog
        try:
            with open(AWDIT_CATALOG_PATH, "w") as f:
                json.dump(catalog, f, indent=2)
            log(f"[Discovery] AWD-IT: catalog saved ({len(catalog)} entries).")
        except Exception as e:
            log(f"[Discovery] AWD-IT: catalog save error: {e}")

        # Try matching again with fresh catalog
        new_matched, final_unmatched = try_match(still_needs)
        matched.update(new_matched)
        if final_unmatched:
            log(f"[Discovery] AWD-IT: {len(final_unmatched)} products still unmatched:")
            for p in final_unmatched[:10]:
                log(f"  - {p['model_no']}")

    log(f"[Discovery] AWD-IT: matched {len(matched)}/{len(needs)} new products.")
    return _write_discovered_urls("AWD-IT URL", matched)


# ── Scan: Google UK search for one LN code ────────────────────────────────────
def _google_scan_url(page, ln_code):
    """Search Google UK for <LN> site:scan.co.uk; return first product URL or None."""
    search_url = (
        f"https://www.google.co.uk/search"
        f"?q={ln_code}+site%3Ascan.co.uk&hl=en-GB&gl=GB&num=5"
    )
    try:
        page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
        time.sleep(random.uniform(3, 5))
    except Exception as e:
        log(f"  [Scan] Navigation error for {ln_code}: {e}")
        return None

    body = page.inner_text("body")[:400].lower()
    if "before you continue" in body or "i'm not a robot" in body:
        log(f"  [Scan] ⚠️  Google consent/CAPTCHA for {ln_code}")
        return None

    urls = page.evaluate(r"""() => {
        const found = [];
        for (const a of document.querySelectorAll('a[href]')) {
            const h = a.href || '';
            const m = h.match(/[?&]q=(https?:\/\/(?:www\.)?scan\.co\.uk\/products\/[^&]+)/);
            if (m) { found.push(decodeURIComponent(m[1])); continue; }
            if (/scan\.co\.uk\/products\//.test(h) && !h.includes('google.'))
                found.push(h);
        }
        return [...new Set(found)];
    }""")

    if urls:
        ln_lower = ln_code.lower()
        for u in urls:
            if ln_lower in u.lower():
                return u
        return urls[0]

    # Regex fallback on raw HTML
    raw = page.content()
    matches = re.findall(r'https?://(?:www\.)?scan\.co\.uk/products/[^\s"\'<>]+', raw)
    if matches:
        cleaned = [re.sub(r'["\'>]+$', '', m) for m in matches if m.startswith("http")]
        if cleaned:
            return cleaned[0]
    return None


# ── Scan URL discovery ────────────────────────────────────────────────────────
def discover_scan_urls(page):
    """
    For every product that has a Scan LN code but no Scan URL, search Google UK
    and write the discovered URL to the 'Scan URL' column.
    """
    needs = load_products_needing_scan_url()
    if not needs:
        log("[Discovery] Scan: all LN codes already have URLs — skipping.")
        return 0

    log(f"[Discovery] Scan: {len(needs)} products need URLs (searching Google UK).")
    found_urls = {}
    failed     = []

    for i, row in enumerate(needs, 1):
        ln    = row["ln_code"]
        pid   = row["product_id"]
        model = row["model_no"]
        log(f"  [Scan] [{i}/{len(needs)}] {ln}  ({model[:40]})")

        url = _google_scan_url(page, ln)
        if url:
            log(f"  [Scan] ✅ {url}")
            found_urls[pid] = url
        else:
            log(f"  [Scan] ❌ not found")
            failed.append(ln)

        if i < len(needs):
            time.sleep(random.uniform(GOOGLE_DELAY_MIN, GOOGLE_DELAY_MAX))

    log(f"[Discovery] Scan: found {len(found_urls)}/{len(needs)} URLs.")
    if failed:
        log(f"[Discovery] Scan: unresolved LN codes: {', '.join(failed[:20])}")
    return _write_discovered_urls("Scan URL", found_urls)


# ── Combined pre-flight orchestrator ─────────────────────────────────────────
def _google_very_url(page, sku):
    """Search Google UK for <sku> site:very.co.uk; return first .prd product URL or None."""
    search_url = (
        f"https://www.google.co.uk/search"
        f"?q={sku}+site%3Avery.co.uk&hl=en-GB&gl=GB&num=5"
    )
    try:
        page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
        time.sleep(random.uniform(3, 5))
    except Exception as e:
        log(f"  [Very] Navigation error for {sku}: {e}")
        return None

    body = page.inner_text("body")[:400].lower()
    if "before you continue" in body or "i'm not a robot" in body:
        log(f"  [Very] ⚠️  Google consent/CAPTCHA for {sku}")
        return None

    urls = page.evaluate(r"""() => {
        const found = [];
        for (const a of document.querySelectorAll('a[href]')) {
            const h = a.href || '';
            const m = h.match(/[?&]q=(https?:\/\/(?:www\.)?very\.co\.uk\/[^&]+\.prd[^&]*)/);
            if (m) { found.push(decodeURIComponent(m[1])); continue; }
            if (/very\.co\.uk\/.*\.prd/.test(h) && !h.includes('google.'))
                found.push(h);
        }
        return [...new Set(found)];
    }""")

    if urls:
        return urls[0]

    raw = page.content()
    matches = re.findall(r'https?://(?:www\.)?very\.co\.uk/[^\s"\'<>]+\.prd', raw)
    if matches:
        cleaned = [re.sub(r'["\'>]+$', '', m) for m in matches]
        if cleaned:
            return cleaned[0]
    return None


def discover_very_urls(page):
    """Discover Very product URLs for products that have a Very SKU but no Very URL."""
    needs = load_products_needing_very_url()
    if not needs:
        log("[Discovery] Very: all SKUs already have URLs — skipping.")
        return 0
    log(f"[Discovery] Very: {len(needs)} products need URLs.")
    found_urls, failed = {}, []
    for i, p in enumerate(needs, 1):
        sku   = p["sku"]
        model = p["model_no"]
        log(f"  [Very] [{i}/{len(needs)}] {sku}  ({model[:40]})")
        time.sleep(random.uniform(GOOGLE_DELAY_MIN, GOOGLE_DELAY_MAX))
        url = _google_very_url(page, sku)
        if url:
            found_urls[p["product_id"]] = url
            log(f"  [Very] ✅ {url}")
        else:
            failed.append(sku)
            log(f"  [Very] ❌ not found")
    log(f"[Discovery] Very: found {len(found_urls)}/{len(needs)} URLs.")
    if failed:
        log(f"[Discovery] Very: unresolved SKUs: {', '.join(failed[:20])}")
    return _write_discovered_urls("Very URL", found_urls)


def run_preflight_discovery():
    """
    Run AWD-IT and Scan URL discovery for any new products added since the last
    run.  Uses its own browser instance so discovery and scraping are isolated.
    Syncs the updated template to OneDrive if any URLs were written.
    """
    log("\n" + "=" * 65)
    log("PRE-FLIGHT URL DISCOVERY")
    log("=" * 65)

    # Quick check: is there anything to do at all?
    awdit_needs = load_products_needing_awdit()
    scan_needs  = load_products_needing_scan_url()
    very_needs  = load_products_needing_very_url()

    if not awdit_needs and not scan_needs and not very_needs:
        log("[Discovery] Nothing to discover — all products have URLs. Skipping.")
        return

    log(f"[Discovery] AWD-IT needs: {len(awdit_needs)} | Scan needs: {len(scan_needs)} | Very needs: {len(very_needs)}")

    total_written = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled",
                  "--disable-dev-shm-usage"]
        )
        ctx = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-GB",
            timezone_id="Europe/London",
            extra_http_headers={
                "Accept-Language": "en-GB,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "DNT": "1",
            }
        )
        ctx.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-GB', 'en']});
            Object.defineProperty(navigator, 'plugins',   {get: () => [1, 2, 3]});
            window.chrome = {runtime: {}};
        """)
        page = ctx.new_page()

        if awdit_needs:
            written = discover_awdit_urls(page)
            total_written += written

        if scan_needs:
            written = discover_scan_urls(page)
            total_written += written

        if very_needs:
            written = discover_very_urls(page)
            total_written += written

        browser.close()

    if total_written:
        log(f"[Discovery] {total_written} new URLs written — syncing template to OneDrive.")
        sync_template_to_onedrive()
    else:
        log("[Discovery] No new URLs found this run.")

    log("PRE-FLIGHT COMPLETE\n" + "=" * 65 + "\n")


# ── Main run ──────────────────────────────────────────────────────────────────
def run(start, end, date_str, notify=False):
    log(f"Starting retailer scrape: rows {start}–{end}, date={date_str}")
    ensure_output_file(date_str)

    id_codes  = load_retailer_ids()
    products  = read_products(start, end)
    completed = load_progress(date_str)

    active  = [r for r in RETAILERS if not r["blocked"]]
    blocked = [r for r in RETAILERS if r["blocked"]]
    log(f"Loaded {len(products)} products | ID codes: {len(id_codes)} | Active: {[r['name'] for r in active]}")

    found = 0
    not_stocked = 0
    not_found = 0
    products_done = 0
    since_pause   = 0
    long_pause_every = random.randint(20, 35)

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled","--disable-dev-shm-usage"]
        )
        context = browser.new_context(
            viewport={"width":1366,"height":768},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="en-GB", timezone_id="Europe/London",
            extra_http_headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
                "Accept-Language": "en-GB,en;q=0.9",
                "DNT": "1",
            }
        )
        context.add_init_script("""
            Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
            Object.defineProperty(navigator,'languages',{get:()=>['en-GB','en']});
            Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});
            window.chrome={runtime:{}};
        """)
        page = context.new_page()

        for product in products:
            row_num   = product["row_num"]
            model_no  = product["model_no"]
            msrp      = product["msrp"]
            excel_row = row_num + 1

            if str(row_num) in completed:
                log(f"Skipping row {row_num} ({model_no}) — already done.")
                continue

            msrp_str = f"£{msrp:.2f}" if msrp else "no MSRP"
            log(f"\n[{row_num}/{end}] {model_no} ({msrp_str})")

            page.mouse.move(random.randint(200,1100), random.randint(150,600))

            # Write blocked retailer cells instantly
            for r in blocked:
                write_cell(date_str, excel_row, r["col"], None, msrp, status="BLOCKED")

            # Scrape active retailers
            for retailer in active:
                name  = retailer["name"]
                price, status = scrape_product(page, product, retailer, id_codes)

                if price is not None:
                    below = " ⚠️ BELOW MSRP" if msrp and price < msrp else ""
                    log(f"    [{name}] £{price:.2f}{below}")
                    write_cell(date_str, excel_row, retailer["col"], price, msrp)
                    found += 1
                elif status == "OUT_OF_STOCK":
                    log(f"    [{name}] out of stock")
                    write_cell(date_str, excel_row, retailer["col"], None, msrp, status="OUT_OF_STOCK")
                    not_stocked += 1
                elif status == "NOT_STOCKED":
                    log(f"    [{name}] not stocked (no ID code)")
                    write_cell(date_str, excel_row, retailer["col"], None, msrp)
                    not_stocked += 1
                else:
                    log(f"    [{name}] not found ({status})")
                    write_cell(date_str, excel_row, retailer["col"], None, msrp)
                    not_found += 1

                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

            products_done += 1
            completed.add(str(row_num))
            save_progress(date_str, completed)

            since_pause += 1
            if since_pause >= long_pause_every:
                pause = random.uniform(LONG_PAUSE_MIN, LONG_PAUSE_MAX)
                log(f"Long pause: {pause:.0f}s")
                time.sleep(pause)
                since_pause = 0
                long_pause_every = random.randint(20, 35)

        browser.close()

    sync_ok = sync_onedrive()
    log(f"\nBatch complete. Products={products_done}, Found={found}, Not stocked={not_stocked}, Not found={not_found}")

    if notify:
        total_today = len(load_progress(date_str))
        icon = "✅" if sync_ok else "⚠️"
        send_telegram(
            f"🛒 <b>Retailer Tracker complete</b> — {date_str}\n\n"
            f"✔️ Total today: {total_today} products\n"
            f"💰 Prices found: {found}\n"
            f"🚫 Not stocked: {not_stocked}\n"
            f"❌ Not found: {not_found}\n"
            f"{icon} OneDrive: {'synced' if sync_ok else 'FAILED'}\n"
            f"📁 Retailer_Results_{datetime.now().strftime('%Y-%m')}.xlsx"
        )

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch",    type=int, choices=[1, 2, 3],
                        help="Run one of three equal-sized batches (discovery runs before batch 1)")
    parser.add_argument("--test",     action="store_true",
                        help="Scrape first 20 products (includes pre-flight discovery)")
    parser.add_argument("--discover", action="store_true",
                        help="Run only the pre-flight URL discovery, then exit")
    parser.add_argument("--start",    type=int, help="Start row (manual range)")
    parser.add_argument("--end",      type=int, help="End row (manual range)")
    args = parser.parse_args()

    date_str = datetime.now().strftime("%d-%m-%Y")

    if args.discover:
        # Standalone discovery run — useful to trigger manually after adding LN/ASIN codes
        sync_template_from_onedrive()
        run_preflight_discovery()

    elif args.test:
        # Test mode: pull latest template, run discovery, then scrape first 20 rows
        sync_template_from_onedrive()
        run_preflight_discovery()
        run(1, 20, date_str, notify=False)

    elif args.batch in (1, 2, 3):
        total  = count_products()
        ranges = batch_ranges(total)
        log(f"Template: {total} products — batch ranges: {ranges}")
        s, e = ranges[args.batch - 1]

        if args.batch == 1:
            # Pull latest template first (user may have updated codes at work via OneDrive)
            # then run URL discovery for any new products before price scraping starts
            sync_template_from_onedrive()
            run_preflight_discovery()
            # Re-read product count after discovery may have updated the file
            total  = count_products()
            ranges = batch_ranges(total)
            s, e   = ranges[0]

        run(s, e, date_str, notify=(args.batch == 3))

    elif args.start and args.end:
        run(args.start, args.end, date_str, notify=False)

    else:
        parser.print_help()
