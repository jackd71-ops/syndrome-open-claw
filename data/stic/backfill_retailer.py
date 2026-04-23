#!/usr/bin/env python3
"""
Backfill retailer_prices table from Retailer_Results_2026-04.xlsx.
Reads every dated sheet, inserts one row per product×retailer.
Safe to re-run — uses INSERT OR IGNORE.

Column mapping (0-based):
  A=0  product_id
  B=1  description
  C=2  model_no
  D=3  manufacturer
  E=4  product_group
  H=7  MSRP
  J=9  Amazon UK
  K=10 Currys
  L=11 Argos
  M=12 Scan
  N=13 Overclockers
  O=14 Box
  P=15 CCL Online
  Q=16 AWD-IT
  R=17 Very

Usage: python3 backfill_retailer.py
"""
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = "/opt/openclaw/data/general/Retailer_Results_2026-04.xlsx"
DB_PATH   = "/opt/openclaw/data/analytics/prices.db"

RETAILER_COLS = [
    ("Amazon",       9),
    ("Currys",       10),
    ("Argos",        11),
    ("Scan",         12),
    ("Overclockers", 13),
    ("Box",          14),
    ("CCL Online",   15),
    ("AWD-IT",       16),
    ("Very",         17),
]

DATE_RE = re.compile(r"^(\d{2})-(\d{2})-(\d{4})$")

_BAD_STRINGS = {"search_failed", "cf_block", "n/a", "—", "-", "oos", ""}


def sheet_to_iso(name: str):
    m = DATE_RE.match(name)
    if not m:
        return None
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"


def to_float(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip().lower() in _BAD_STRINGS:
        return None
    if isinstance(v, str):
        v = v.replace("£", "").replace(",", "").strip()
    try:
        f = float(v)
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def backfill():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")

    total_inserted = 0
    total_skipped  = 0

    for sheet_name in wb.sheetnames:
        iso_date = sheet_to_iso(sheet_name)
        if not iso_date:
            print(f"  Skipping sheet '{sheet_name}' (not a date)")
            continue

        ws = wb[sheet_name]
        rows_inserted = 0
        rows_skipped  = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            product_id = row[0]
            if product_id is None:
                continue
            try:
                product_id = int(product_id)
            except (ValueError, TypeError):
                continue

            description   = str(row[1]).strip() if row[1] else ""
            model_no      = str(row[2]).strip() if row[2] else ""
            manufacturer  = str(row[3]).strip() if row[3] else ""
            product_group = str(row[4]).strip() if row[4] else None
            msrp          = to_float(row[7])

            for retailer, col_idx in RETAILER_COLS:
                price = to_float(row[col_idx]) if col_idx < len(row) else None

                below_msrp = None
                if price is not None and msrp is not None:
                    below_msrp = 1 if price < msrp else 0

                cur = db.execute(
                    """INSERT OR IGNORE INTO retailer_prices
                       (date, product_id, model_no, description, manufacturer,
                        product_group, msrp, retailer, price, below_msrp)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (iso_date, product_id, model_no, description, manufacturer,
                     product_group, msrp, retailer, price, below_msrp),
                )
                if cur.rowcount:
                    rows_inserted += 1
                else:
                    rows_skipped += 1

        db.commit()
        total_inserted += rows_inserted
        total_skipped  += rows_skipped
        print(f"  {sheet_name} → {iso_date}: inserted {rows_inserted}, skipped {rows_skipped}")

    db.close()
    wb.close()
    print(f"\nDone. Total inserted: {total_inserted}, skipped: {total_skipped}")


if __name__ == "__main__":
    backfill()
