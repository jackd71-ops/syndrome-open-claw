#!/usr/bin/env python3
"""
Nightly STIC template sync — runs at 00:00 UK time (01:00 BST in summer).

Bidirectional sync between OneDrive Excel and the products SQLite table:
  PULL  OneDrive → local Excel
  READ  Excel    → upsert products table (new SKUs, EAN, chipset, etc.)
  READ  DB       → write EOL flags back into Excel col H  (batches portal toggles)
  PUSH  local Excel → OneDrive  (one rclone, once per night)
"""

import sqlite3
import subprocess
import sys
from datetime import datetime
from pathlib import Path

DB_PATH      = "/opt/openclaw/data/analytics/prices.db"
MASTER_PATH  = "/opt/openclaw/data/general/STIC Template.xlsx"
ONEDRIVE_SRC = "onedrive:Documents/STIC/STIC Template.xlsx"
LOG_PATH     = "/opt/openclaw/logs/sync_template.log"


def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_PATH, "a") as f:
        f.write(line + "\n")


def ensure_table(db):
    db.execute("""
        CREATE TABLE IF NOT EXISTS products (
            product_id    INTEGER PRIMARY KEY,
            description   TEXT,
            model_no      TEXT,
            manufacturer  TEXT,
            product_group TEXT,
            chipset       TEXT,
            ean           TEXT,
            eol           INTEGER NOT NULL DEFAULT 0,
            stic_url      TEXT
        )
    """)
    try:
        db.execute("ALTER TABLE products ADD COLUMN stic_url TEXT")
    except Exception:
        pass  # column already exists
    db.commit()


def pull_from_onedrive() -> bool:
    log("Pulling template from OneDrive...")
    try:
        result = subprocess.run(
            ["rclone", "copyto", ONEDRIVE_SRC, MASTER_PATH],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            log("Pull OK.")
            return True
        log(f"Pull WARNING (using local copy): {result.stderr.strip()}")
        return False
    except subprocess.TimeoutExpired:
        log("Pull TIMEOUT — using local copy.")
        return False


def push_to_onedrive() -> bool:
    log("Pushing updated template to OneDrive...")
    try:
        result = subprocess.run(
            ["rclone", "copyto", MASTER_PATH, ONEDRIVE_SRC],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            log("Push OK.")
            return True
        log(f"Push FAILED: {result.stderr.strip()}")
        return False
    except subprocess.TimeoutExpired:
        log("Push TIMEOUT — will retry on next nightly run.")
        return False


def sync():
    from openpyxl import load_workbook

    pull_from_onedrive()

    if not Path(MASTER_PATH).exists():
        log("ERROR: template file not found — aborting.")
        sys.exit(1)

    log("Reading template...")
    wb = load_workbook(MASTER_PATH, read_only=True)
    ws = wb.worksheets[0]

    excel_rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        excel_rows.append({
            "product_id":    int(row[0]),
            "description":   str(row[1]).strip() if row[1] else None,
            "model_no":      str(row[2]).strip() if row[2] else None,
            "manufacturer":  str(row[3]).strip() if row[3] else None,
            "product_group": str(row[4]).strip() if row[4] else None,
            "chipset":       str(row[5]).strip() if len(row) > 5 and row[5] else None,
            "ean":           str(row[6]).strip() if len(row) > 6 and row[6] else None,
            "eol_excel":     bool(row[7])         if len(row) > 7 and row[7] else False,
        })
    wb.close()
    log(f"Read {len(excel_rows)} products from Excel.")

    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    ensure_table(db)

    existing = {r["product_id"] for r in db.execute("SELECT product_id FROM products").fetchall()}

    upserted = 0
    new_count = 0
    for p in excel_rows:
        pid = p["product_id"]
        if pid in existing:
            db.execute("""
                UPDATE products SET
                    description=?, model_no=?, manufacturer=?,
                    product_group=?, chipset=?, ean=?
                WHERE product_id=?
            """, (p["description"], p["model_no"], p["manufacturer"],
                  p["product_group"], p["chipset"], p["ean"], pid))
        else:
            db.execute("""
                INSERT INTO products
                    (product_id, description, model_no, manufacturer,
                     product_group, chipset, ean, eol)
                VALUES (?,?,?,?,?,?,?,?)
            """, (pid, p["description"], p["model_no"], p["manufacturer"],
                  p["product_group"], p["chipset"], p["ean"],
                  1 if p["eol_excel"] else 0))
            new_count += 1
        upserted += 1

    db.commit()
    log(f"Upserted {upserted} products ({new_count} new).")

    eol_db = {r["product_id"]: r["eol"]
              for r in db.execute("SELECT product_id, eol FROM products").fetchall()}
    db.close()

    wb_write = load_workbook(MASTER_PATH)
    ws_write = wb_write.worksheets[0]

    eol_written = 0
    for row in ws_write.iter_rows(min_row=3):
        pid = row[0].value
        if pid is None:
            continue
        pid = int(pid)
        db_eol = eol_db.get(pid, 0)
        current_excel_eol = row[7].value if len(row) > 7 else None
        new_val = "Y" if db_eol else None
        if new_val != current_excel_eol:
            row[7].value = new_val
            eol_written += 1

    log(f"EOL flags written back to Excel: {eol_written} changes.")
    wb_write.save(MASTER_PATH)
    wb_write.close()

    push_to_onedrive()

    db2 = sqlite3.connect(DB_PATH)
    total   = db2.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    active  = db2.execute("SELECT COUNT(*) FROM products WHERE eol=0").fetchone()[0]
    eol_tot = db2.execute("SELECT COUNT(*) FROM products WHERE eol=1").fetchone()[0]
    db2.close()
    log(f"Sync complete. DB: {total} total, {active} active, {eol_tot} EOL.")


if __name__ == "__main__":
    sync()
