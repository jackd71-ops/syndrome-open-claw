#!/usr/bin/env python3
"""
One-off seed: populate retailer_ids and products.msrp from Retailer_Template.xlsx.
Run once after DB schema migration; safe to re-run (uses INSERT OR REPLACE).
"""
import sqlite3
from openpyxl import load_workbook

TEMPLATE_PATH = "/opt/openclaw/data/general/Retailer_Template.xlsx"
DB_PATH       = "/opt/openclaw/data/analytics/prices.db"

# Excel header → DB column mapping (normalise away the double-space in ARGOS)
HEADER_MAP = {
    "Amazon ASIN": "amazon_asin",
    "Currys SKU":  "currys_sku",
    "Very SKU":    "very_sku",
    "ARGOS  SKU":  "argos_sku",
    "ARGOS SKU":   "argos_sku",
    "CCL URL":     "ccl_url",
    "AWD-IT URL":  "awdit_url",
    "Scan LN":     "scan_ln",
    "Scan URL":    "scan_url",
    "OCUK Code":   "ocuk_code",
    "Box URL":     "box_url",
}

def seed():
    wb = load_workbook(TEMPLATE_PATH, read_only=True)

    ws_ids   = wb["Retailer_IDs"]
    raw_hdrs = [c.value for c in ws_ids[1]]
    col_map  = {}
    for i, hdr in enumerate(raw_hdrs):
        if hdr and i >= 4:
            db_col = HEADER_MAP.get(str(hdr).strip())
            if db_col:
                col_map[i] = db_col

    print(f"Retailer_IDs col map: {col_map}")

    retailer_rows = []
    for row in ws_ids.iter_rows(min_row=2, values_only=True):
        pid = row[0]
        if pid is None:
            continue
        try:
            pid = int(str(pid).strip())
        except ValueError:
            continue
        entry = {"product_id": pid}
        for i, db_col in col_map.items():
            val = row[i] if i < len(row) else None
            if val is not None:
                v = str(val).strip()
                if v and v.lower() != "none":
                    entry[db_col] = v
        retailer_rows.append(entry)

    ws_master = wb.worksheets[0]
    msrp_map  = {}
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        pid = row[0]
        if pid is None:
            continue
        try:
            pid = int(str(pid).strip())
        except ValueError:
            continue
        msrp = None
        try:
            msrp = float(row[7]) if row[7] is not None else None
        except (ValueError, TypeError):
            pass
        if msrp:
            msrp_map[pid] = msrp

    wb.close()

    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA journal_mode=WAL")

    inserted = 0
    for entry in retailer_rows:
        pid = entry.pop("product_id")
        if not entry:
            con.execute("INSERT OR IGNORE INTO retailer_ids (product_id) VALUES (?)", (pid,))
        else:
            cols         = ", ".join(["product_id"] + list(entry.keys()))
            placeholders = ", ".join(["?"] * (1 + len(entry)))
            vals         = [pid] + list(entry.values())
            con.execute(
                f"INSERT OR REPLACE INTO retailer_ids ({cols}) VALUES ({placeholders})", vals
            )
        inserted += 1

    print(f"retailer_ids: {inserted} rows inserted/replaced")

    updated = 0
    for pid, msrp in msrp_map.items():
        cur = con.execute(
            "UPDATE products SET msrp=? WHERE product_id=? AND msrp IS NULL", (msrp, pid)
        )
        updated += cur.rowcount

    print(f"products.msrp: {updated} rows updated")
    con.commit()
    con.close()
    print("Seed complete.")

if __name__ == "__main__":
    seed()
